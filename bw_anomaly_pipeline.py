
import os, re, shutil, datetime as dt
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import warnings
warnings.filterwarnings('ignore')

CONFIG = {
    "base_dir": ".",
    "materiality_vnd": 1_000_000_000,
    "recurring_pct_threshold": 0.05,
    "revenue_opex_pct_threshold": 0.10,
    "bs_pct_threshold": 0.05,
    "archive_processed": True,
    "recurring_code_prefixes": ["6321", "635", "515"],
    "min_trend_periods": 3
}

MONTHS = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]
BS_PAT = re.compile(r'^\s*as\s*of\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[\.\-\s]*(\d{2,4})\s*$', re.I)
PL_PAT = re.compile(r'^\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[\.\-\s]*(\d{2,4})\s*$', re.I)

def normalize_period_label(label):
    if label is None: return ""
    s = str(label).strip()
    if s == "": return ""
    try:
        s_clean = re.sub(r'^\s*(as\s*of|tinh\s*den|tính\s*đến|den\s*ngay|đến\s*ngày)\s*', '', s, flags=re.I)
        m = re.search(r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[^\w]?[\s\-\.]*([12]\d{3}|\d{2})\b', s_clean, flags=re.I)
        if m:
            mon, yr = m.group(1), m.group(2); yr = int(yr); yr = yr+2000 if yr < 100 else yr
            return f"{mon.title()} {yr}"
        m = re.search(r'\b(1[0-2]|0?[1-9])[./\-](\d{4})\b', s_clean)
        if m: mon=int(m.group(1)); yr=int(m.group(2)); return f"{MONTHS[mon-1].title()} {yr}"
        m = re.search(r'\b(\d{4})[./\-](1[0-2]|0?[1-9])\b', s_clean)
        if m: yr=int(m.group(1)); mon=int(m.group(2)); return f"{MONTHS[mon-1].title()} {yr}"
        m_year = re.search(r'(20\d{2}|19\d{2})', s_clean); m_mon = re.search(r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\b', s_clean, flags=re.I)
        if m_year and m_mon: yr=int(m_year.group(1)); mon=m_mon.group(0); return f"{mon.title()} {yr}"
    except Exception: pass
    return s

def month_key(label):
    n = normalize_period_label(label)
    m = re.search(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+(\d{4})', n, re.I)
    if not m: return (9999, 99)
    y = int(m.group(2)); mi = MONTHS.index(m.group(1).lower())+1
    return (y, mi)

def detect_header_row(xl, sheet):
    try:
        probe = pd.read_excel(xl, sheet_name=sheet, header=None, nrows=40)
        for i in range(len(probe)):
            row_values = probe.iloc[i].astype(str).str.strip().str.lower()
            if any("financial row" in v for v in row_values):
                return i
    except Exception as e:
        print(f"Warning: header detect in {sheet}: {e}")
    return 0

def normalize_financial_col(df):
    for c in df.columns:
        if str(c).strip().lower() == "financial row":
            return df.rename(columns={c: "Financial row"})
    return df.rename(columns={df.columns[0]: "Financial row"})

def promote_row8(df, mode, sub):
    if len(df) < 1: return df, []
    row8 = df.iloc[0]
    new_cols = []
    for c in df.columns:
        v = str(row8.get(c, "")).strip()
        if BS_PAT.match(v) or PL_PAT.match(v):
            new_cols.append(normalize_period_label(v))
        else:
            new_cols.append(str(c))
    df = df.copy(); df.columns = new_cols; df = df.iloc[1:].reset_index(drop=True)
    month_cols = []
    for c in df.columns:
        normalized = normalize_period_label(c)
        if re.match(r'^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+\d{4}$', normalized, re.I):
            month_cols.append(c)
    month_cols = sorted(month_cols, key=month_key)
    print(f"[{sub}] {mode} periods detected: {month_cols}")
    return df, month_cols

def fill_down_assign(df):
    """
    Tag rows; keep header/detail/total rows for smart aggregation:
      - RowHadOwnCode: header rows (code present in col A)
      - IsTotal: 'Total ...' with a code
    """
    ser = df["Financial row"].astype(str)
    code_extract = ser.str.extract(r'(\d{4,})', expand=False)
    name_extract = ser.str.replace(r'.*?(\d{4,})\s*[-:]*\s*', '', regex=True).str.strip()

    row_has_code = code_extract.notna()
    is_total_word = ser.str.strip().str.lower().str.startswith(("total","subtotal","cộng","tong","tổng"))
    is_total_with_code = is_total_word & row_has_code
    is_section = ser.str.match(r'^\s*([IVX]+\.|[A-Z]\.)\s')
    is_empty = ser.str.strip().eq("")

    df["Account Code"] = code_extract.ffill()
    df["Account Name"] = name_extract.where(row_has_code).ffill()
    df["RowHadOwnCode"] = row_has_code
    df["IsTotal"] = is_total_with_code

    keep_mask = ~(is_section | is_empty)
    df = df[keep_mask & df["Account Code"].notna()].copy()
    return df

def coerce_numeric(df, month_cols):
    out = df.copy()
    for c in month_cols:
        if c in out.columns:
            series = out[c].astype(str)
            series = (series
                .str.replace("\u00a0","", regex=False)
                .str.replace(",","", regex=False)
                .str.replace(r"\((.*)\)", r"-\1", regex=True)
                .str.replace(r"[^0-9\.\-]", "", regex=True)
            )
            out[c] = pd.to_numeric(series, errors="coerce").fillna(0.0)
    return out

def aggregate_totals(df, month_cols):
    """Smart aggregation per account:
       - Use 'Total' rows when present for a code;
       - Otherwise sum all rows of that code (header + detail).
    """
    if df.empty:
        return pd.DataFrame(columns=["Account Code","Account Name"] + month_cols)

    if "RowHadOwnCode" in df.columns:
        nm_src = df[df["RowHadOwnCode"]]
    else:
        nm_src = df
    name_map = nm_src.dropna(subset=["Account Code"])[["Account Code","Account Name"]]\
                     .drop_duplicates("Account Code")\
                     .set_index("Account Code")["Account Name"]

    totals_df = df[df.get("IsTotal", False)]
    codes_with_total = set(totals_df["Account Code"].dropna().astype(str).unique())

    cols = ["Account Code"] + [c for c in month_cols if c in df.columns]
    parts = []
    if not totals_df.empty:
        parts.append(totals_df[cols].groupby("Account Code", as_index=False).sum())
    no_total_df = df[~df["Account Code"].astype(str).isin(codes_with_total)]
    if not no_total_df.empty:
        parts.append(no_total_df[cols].groupby("Account Code", as_index=False).sum())
    agg = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame(columns=cols)
    agg["Account Name"] = agg["Account Code"].map(name_map).fillna("")
    return agg[["Account Code","Account Name"] + [c for c in month_cols if c in agg.columns]]

def compute_mom_with_trends(df, month_cols):
    if len(month_cols) < 2:
        return pd.DataFrame(columns=["Account Code","Account Name","Prior","Current","Delta","Pct Change","Period","Trend_3M_Avg","Trend_Deviation"])
    out = []
    for i in range(1, len(month_cols)):
        cur, prev = month_cols[i], month_cols[i-1]
        if cur not in df.columns or prev not in df.columns: continue
        tmp = df[["Account Code","Account Name", prev, cur]].copy()
        tmp = tmp.rename(columns={prev: "Prior", cur: "Current"})
        tmp["Delta"] = tmp["Current"] - tmp["Prior"]
        tmp["Pct Change"] = np.where(tmp["Prior"] == 0, np.nan, tmp["Delta"] / tmp["Prior"])
        tmp["Period"] = normalize_period_label(cur)
        if i >= CONFIG["min_trend_periods"]:
            start_idx = max(0, i - 5)
            trend_cols = month_cols[start_idx:i]
            if len(trend_cols) >= CONFIG["min_trend_periods"]:
                trend_data = df[trend_cols]
                tmp["Trend_3M_Avg"] = trend_data.mean(axis=1)
                tmp["Trend_Deviation"] = tmp["Current"] - tmp["Trend_3M_Avg"]
            else:
                tmp["Trend_3M_Avg"] = np.nan; tmp["Trend_Deviation"] = np.nan
        else:
            tmp["Trend_3M_Avg"] = np.nan; tmp["Trend_Deviation"] = np.nan
        out.append(tmp)
    return pd.concat(out, ignore_index=True) if out else pd.DataFrame(columns=["Account Code","Account Name","Prior","Current","Delta","Pct Change","Period","Trend_3M_Avg","Trend_Deviation"])

def classify_pl_account(code):
    code_str = str(code)
    return "Recurring" if any(code_str.startswith(p) for p in CONFIG["recurring_code_prefixes"]) else "Revenue/OPEX"

def get_threshold_cause(statement, code):
    if statement == "BS": return "Balance changed materially — check reclass/missing offset."
    return "Recurring moved — check accruals/timing." if classify_pl_account(code) == "Recurring" else "Revenue/OPEX moved — check billing/cut-off."

def match_codes(series, pattern_str):
    if pd.isna(pattern_str) or pattern_str == "":
        return pd.Series(False, index=series.index)
    patterns = [p.strip() for p in str(pattern_str).split("|") if p.strip()]
    mask = pd.Series(False, index=series.index)
    for pattern in patterns:
        if pattern.endswith("*"):
            prefix = pattern[:-1]
            mask |= series.astype(str).str.startswith(prefix)
        else:
            mask |= (series.astype(str) == pattern)
    return mask

def build_corr_anoms(sub,combined,corr_rules,periods,materiality):
    items=[]
    cols = {c.lower(): c for c in corr_rules.columns}
    def pick(opts):
        for n in opts:
            if n in cols: return cols[n]
        return None
    left_col  = pick(["left_patterns","left_pattern","left_patter","left"])
    right_col = pick(["right_patterns","right_pattern","right_patter","right"])
    cause_col = pick(["cause_message","cause","message","notes"])
    type_col  = pick(["relation_type","type","direction"])
    for _,rule in corr_rules.iterrows():
        lp = str(rule[left_col]) if left_col else ""
        rp = str(rule[right_col]) if right_col else ""
        cause = str(rule[cause_col]) if cause_col else "Correlation mismatch"
        rel = str(rule.get(type_col,"directional")).strip().lower() if type_col else "directional"
        inverse = rel in ("inverse","opposite","neg","negative")
        if not lp or not rp: 
            continue
        for per in periods:
            mom=combined[combined["Norm_Period"]==per]
            if mom.empty: continue
            l=mom[match_codes(mom["Account Code"],lp)]["Delta"].sum()
            r=mom[match_codes(mom["Account Code"],rp)]["Delta"].sum()
            ok = ((l>0 and r<0) or (l<0 and r>0)) if inverse else ((l>0 and r>0) or (l<0 and r<0))
            if abs(l)>=materiality and not ok:
                items.append({
                    "Subsidiary":sub,"Account":f"{lp} ↔ {rp}","Period":per,"Pct Change":"",
                    "Abs Change (VND)":int(l),"Trigger(s)":"Correlation break",
                    "Suggested likely cause":cause,"Status":"Needs Review",
                    "Notes":f"Left Δ={int(l):,}, Right Δ={int(r):,}, relation={'inverse' if inverse else 'directional'}"
                })
    return items

def build_anoms(sub, bs_data, bs_cols, pl_data, pl_cols, corr_rules, season_rules):
    anomalies = []
    materiality = CONFIG["materiality_vnd"]
    bs_mom = compute_mom_with_trends(bs_data, bs_cols)
    pl_mom = compute_mom_with_trends(pl_data, pl_cols)

    for _, row in bs_mom.iterrows():
        abs_delta = abs(row["Delta"]); pct_change = row["Pct Change"]
        if (abs_delta >= materiality and pd.notna(pct_change) and abs(pct_change) > CONFIG["bs_pct_threshold"]):
            anomalies.append({
                "Subsidiary": sub,"Account": f'{row["Account Code"]}-{row["Account Name"]}',
                "Period": row["Period"],"Pct Change": round(pct_change * 100, 2),
                "Abs Change (VND)": int(row["Delta"]),"Trigger(s)": "BS >5% & ≥1B",
                "Suggested likely cause": get_threshold_cause("BS", row["Account Code"]),
                "Status": "Needs Review","Notes": ""
            })

    for _, row in pl_mom.iterrows():
        abs_delta = abs(row["Delta"]); pct_change = row["Pct Change"]
        account_class = classify_pl_account(row["Account Code"]); trigger = ""
        if account_class == "Recurring":
            if (abs_delta >= materiality and pd.notna(pct_change) and abs(pct_change) > CONFIG["recurring_pct_threshold"]):
                trigger = "Recurring >5% & ≥1B"
        else:
            if ((pd.notna(pct_change) and abs(pct_change) > CONFIG["revenue_opex_pct_threshold"]) or abs_delta >= materiality):
                trigger = "Revenue/OPEX >10% or ≥1B"
        if trigger:
            anomalies.append({
                "Subsidiary": sub,"Account": f'{row["Account Code"]}-{row["Account Name"]}',
                "Period": row["Period"],"Pct Change": round(row["Pct Change"]*100,2) if pd.notna(row["Pct Change"]) else "",
                "Abs Change (VND)": int(row["Delta"]),"Trigger(s)": trigger,
                "Suggested likely cause": get_threshold_cause("PL", row["Account Code"]),
                "Status": "Needs Review","Notes": ""
            })

    combined = pd.concat([bs_mom[["Account Code","Period","Delta"]], pl_mom[["Account Code","Period","Delta"]]])
    combined["Norm_Period"] = combined["Period"].astype(str).map(normalize_period_label)
    periods = sorted(set(combined["Norm_Period"]), key=month_key)
    anomalies.extend(build_corr_anoms(sub, combined, corr_rules, periods, materiality))

    return pd.DataFrame(anomalies)

def process_financial_tab(xl_file, sheet_name, mode, subsidiary):
    try:
        header_row = detect_header_row(xl_file, sheet_name)
        df = pd.read_excel(xl_file, sheet_name=sheet_name, header=header_row, dtype=str)
        df = normalize_financial_col(df)
        df, month_cols = promote_row8(df, mode, subsidiary)
        df = fill_down_assign(df)
        df = coerce_numeric(df, month_cols)
        keep_cols = ["Account Code","Account Name","RowHadOwnCode","IsTotal"] + [c for c in month_cols if c in df.columns]
        df = df[keep_cols]
        totals = aggregate_totals(df, month_cols)
        return totals, month_cols
    except Exception as e:
        print(f"Error processing {sheet_name} for {subsidiary}: {e}")
        return pd.DataFrame(), []

def extract_subsidiary_name(xl_file):
    try:
        wb = load_workbook(xl_file, read_only=True, data_only=True)
        for sheet_name in ["BS Breakdown", "PL Breakdown"]:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                cell_value = sheet["A2"].value
                if isinstance(cell_value, str) and ":" in cell_value:
                    return cell_value.split(":")[-1].strip()
        wb.close()
    except Exception as e:
        print(f"Warning: Could not extract subsidiary name from {xl_file}: {e}")
    return Path(xl_file).stem.split("_")[0]

def apply_excel_formatting(filepath, anomaly_df):
    try:
        wb = load_workbook(filepath)
        ws = wb["Anomalies Summary"]
        critical_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        warning_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        for cell in ws[1]: cell.fill = header_fill; cell.font = Font(bold=True)
        for row_idx, (_, row) in enumerate(anomaly_df.iterrows(), start=2):
            abs_change = abs(row.get("Abs Change (VND)", 0)); trigger = str(row.get("Trigger(s)", ""))
            fill = None
            if abs_change >= CONFIG["materiality_vnd"] * 5: fill = critical_fill
            elif "Correlation break" in trigger or abs_change >= CONFIG["materiality_vnd"] * 2: fill = warning_fill
            if fill:
                for col_idx in range(1, len(anomaly_df.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill
        wb.save(filepath); wb.close()
    except Exception as e:
        print(f"Warning: Could not apply formatting to {filepath}: {e}")

def main():
    base = Path(CONFIG["base_dir"]).resolve()
    in_dir, out_dir, arc_dir, logic_dir = base/"input", base/"output", base/"archive", base/"logic"
    in_dir.mkdir(exist_ok=True); out_dir.mkdir(exist_ok=True); arc_dir.mkdir(exist_ok=True); logic_dir.mkdir(exist_ok=True)

    mapping_file = logic_dir / "Mapping_ACTIVE.xlsx"
    corr_rules = pd.read_excel(mapping_file, sheet_name="CorrelationRules")
    season_rules = pd.read_excel(mapping_file, sheet_name="Seasonality")

    files = list(in_dir.glob("*.xlsx"))
    print(f"Found {len(files)} Excel files to process")

    all_anoms = []
    for excel_file in files:
        try:
            sub = extract_subsidiary_name(excel_file)
            print(f"\\nProcessing {sub} from {excel_file.name}")
            bs, bs_cols = process_financial_tab(excel_file, "BS Breakdown", "BS", sub)
            pl, pl_cols = process_financial_tab(excel_file, "PL Breakdown", "PL", sub)
            if bs.empty and pl.empty:
                print(f"Warning: No data found for {sub}"); continue
            anoms = build_anoms(sub, bs, bs_cols, pl, pl_cols, corr_rules, season_rules)

            timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            out_file = out_dir / f"Anomalies_{sub}_{timestamp}.xlsx"
            with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
                if not bs.empty: bs.to_excel(writer, sheet_name="BS_cleaned", index=False)
                if not pl.empty: pl.to_excel(writer, sheet_name="PL_cleaned", index=False)
                anoms.to_excel(writer, sheet_name="Anomalies Summary", index=False)
            if not anoms.empty: apply_excel_formatting(out_file, anoms)

            print(f"✓ Generated {out_file.name} with {len(anoms)} anomalies")
            if not anoms.empty: all_anoms.append(anoms.assign(Source_File=excel_file.name))
            if CONFIG["archive_processed"]:
                try: shutil.move(str(excel_file), str(arc_dir / excel_file.name)); print(f"✓ Archived {excel_file.name}")
                except Exception as e: print(f"Warning: Could not archive {excel_file.name}: {e}")
        except Exception as e:
            print(f"Error processing {excel_file.name}: {e}"); continue

    if all_anoms:
        consolidated = pd.concat(all_anoms, ignore_index=True)
        timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        consolidated_file = out_dir / f"Anomalies_CONSOLIDATED_{timestamp}.xlsx"
        with pd.ExcelWriter(consolidated_file, engine="openpyxl") as writer:
            consolidated.to_excel(writer, sheet_name="Anomalies Summary", index=False)
            summary = consolidated.groupby("Subsidiary").agg({"Account":"count","Abs Change (VND)":"sum"}).rename(columns={"Account":"Anomaly_Count","Abs Change (VND)":"Total_Impact_VND"})
            summary.to_excel(writer, sheet_name="Summary by Subsidiary")
        apply_excel_formatting(consolidated_file, consolidated)
        print(f"\\n✓ Generated consolidated report: {consolidated_file.name}")
        print(f"Total anomalies across all subsidiaries: {len(consolidated)}")
    else:
        print("\\nNo anomalies found across all processed files")
    print(f"\\nProcessing complete. Results saved to {out_dir}")

if __name__ == "__main__":
    main()
