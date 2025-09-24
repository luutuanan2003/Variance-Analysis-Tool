# app/main_orchestration.py
"""Main orchestration functions for processing files and generating Excel output."""

from __future__ import annotations

import io
from typing import Optional, List, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from .data_utils import DEFAULT_CONFIG
from .excel_processing import (
    extract_subsidiary_name_from_bytes, process_financial_tab_from_bytes,
    apply_excel_formatting_ws, _add_revenue_analysis_to_sheet
)
from .anomaly_detection import build_anoms_python_mode, build_anoms_ai_mode
from .revenue_analysis import analyze_comprehensive_revenue_impact_from_bytes

def process_all(
    files: list[tuple[str, bytes]],
    corr_rules: Optional[pd.DataFrame] = None,
    season_rules: Optional[pd.DataFrame] = None,
    CONFIG: dict = DEFAULT_CONFIG,
    progress_callback=None
) -> bytes | tuple[bytes, list[tuple[str, bytes]]]:
    """
    Unified processing function that handles both Python and AI analysis modes.

    Args:
        files: List of (filename, bytes) tuples
        corr_rules: Correlation rules DataFrame (Python mode only)
        season_rules: Seasonality rules DataFrame (Python mode only)
        CONFIG: Configuration dictionary
        progress_callback: Progress callback function (AI mode only)

    Returns:
        bytes: Excel file bytes (Python mode)
        tuple[bytes, list[tuple[str, bytes]]]: Excel bytes and debug files (AI mode)
    """
    use_ai = CONFIG.get("use_llm_analysis", False)

    if use_ai:
        return process_all_ai_mode(files, CONFIG, progress_callback)
    else:
        return process_all_python_mode(files, corr_rules, season_rules, CONFIG)

def process_all_python_mode(
    files: list[tuple[str, bytes]],
    corr_rules: Optional[pd.DataFrame] = None,
    season_rules: Optional[pd.DataFrame] = None,
    CONFIG: dict = DEFAULT_CONFIG
) -> bytes:
    """Python rule-based analysis mode."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Anomalies Summary"
    all_anoms: list[pd.DataFrame] = []

    # default empty rules if None
    corr_rules = corr_rules if corr_rules is not None else pd.DataFrame()
    season_rules = season_rules if season_rules is not None else pd.DataFrame()

    for fname, xl_bytes in files:
        sub = extract_subsidiary_name_from_bytes(xl_bytes, fname)

        # Be forgiving if a sheet is missing
        bs_df, bs_cols = pd.DataFrame(), []
        pl_df, pl_cols = pd.DataFrame(), []
        try:
            bs_df, bs_cols = process_financial_tab_from_bytes(xl_bytes, "BS Breakdown", "BS", sub)
        except Exception:
            pass
        try:
            pl_df, pl_cols = process_financial_tab_from_bytes(xl_bytes, "PL Breakdown", "PL", sub)
        except Exception:
            pass

        anoms = build_anoms_python_mode(sub, bs_df, bs_cols, pl_df, pl_cols, corr_rules, season_rules, CONFIG)
        if anoms is not None and not anoms.empty:
            all_anoms.append(anoms)

    # Safe concat (even if no anomalies/files)
    if all_anoms:
        anom_df = pd.concat(all_anoms, ignore_index=True)
    else:
        anom_df = pd.DataFrame(columns=[
            "Subsidiary","Account","Period","Pct Change","Abs Change (VND)",
            "Trigger(s)","Suggested likely cause","Status","Notes"
        ])

    for r in dataframe_to_rows(anom_df, index=False, header=True):
        ws.append(r)
    apply_excel_formatting_ws(ws, anom_df, CONFIG)

    # === ADD REVENUE ANALYSIS SHEET ===
    print(f"\n📊 Adding Revenue Analysis sheet...")
    try:
        # Run revenue analysis for the first file
        if files:
            first_file_name, first_file_bytes = files[0]
            revenue_analysis = analyze_comprehensive_revenue_impact_from_bytes(first_file_bytes, first_file_name)

            # Create revenue analysis sheet
            rev_ws = wb.create_sheet(title="Revenue Analysis")
            _add_revenue_analysis_to_sheet(rev_ws, revenue_analysis)
            print(f"✅ Revenue Analysis sheet added successfully")
    except Exception as e:
        print(f"⚠️  Revenue Analysis sheet creation failed: {e}")
        # Continue without revenue analysis if it fails

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def process_all_ai_mode(
    files: list[tuple[str, bytes]],
    CONFIG: dict = DEFAULT_CONFIG,
    progress_callback=None
) -> tuple[bytes, list[tuple[str, bytes]]]:
    """AI-powered analysis mode."""
    print(f"\n🚀 ===== STARTING AI VARIANCE ANALYSIS PROCESSING =====\n")
    print(f"📥 Processing {len(files)} Excel file(s) for AI analysis")
    print(f"🤖 LLM Model: {CONFIG.get('llm_model', 'gpt-4o')}")
    print(f"🔧 AI-Only Mode: {CONFIG.get('use_llm_analysis', True)}")

    # === EXCEL WORKBOOK INITIALIZATION ===
    print(f"\n📊 Initializing Excel workbook for results...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Anomalies Summary"
    all_anoms: list[pd.DataFrame] = []
    debug_files: list[tuple[str, bytes]] = []  # Store debug files for download
    print(f"✅ Excel workbook initialized successfully")

    # === MULTI-FILE PROCESSING LOOP ===
    print(f"\n🔄 Starting processing loop for {len(files)} file(s)...\n")

    for file_idx, (fname, xl_bytes) in enumerate(files, 1):
        # Calculate progress range for this file (30% to 80% of total)
        file_start = 30 + ((file_idx - 1) * 50 // len(files))
        file_end = 30 + (file_idx * 50 // len(files))

        if progress_callback:
            progress_callback(file_start, f"Processing file {file_idx}/{len(files)}: {fname}")

        print(f"\n📁 ===== PROCESSING FILE {file_idx}/{len(files)} =====\n")
        print(f"📄 File: {fname}")
        print(f"📏 File Size: {len(xl_bytes):,} bytes ({len(xl_bytes)/1024:.1f} KB)")

        if progress_callback:
            progress_callback(file_start + 2, f"Extracting subsidiary name from {fname}")

        print(f"\n🏢 Extracting subsidiary name...")
        sub = extract_subsidiary_name_from_bytes(xl_bytes, fname)
        print(f"✅ Subsidiary: '{sub}'")

        if progress_callback:
            progress_callback(file_start + 5, f"Starting AI analysis for {sub}")

        # === AI ANALYSIS ===
        print(f"\n🤖 Starting AI analysis for '{sub}'...")
        anoms = build_anoms_ai_mode(sub, xl_bytes, fname, CONFIG)

        if progress_callback:
            progress_callback(file_end - 5, f"AI analysis complete for {sub}")

        if anoms is not None and not anoms.empty:
            print(f"✅ AI analysis completed successfully")
            print(f"   • Anomalies detected: {len(anoms)}")
            if len(anoms) > 0:
                ai_status_count = anoms['Status'].value_counts().to_dict()
                for status, count in ai_status_count.items():
                    print(f"   • {status}: {count}")
            all_anoms.append(anoms)
        else:
            print(f"⚠️  No anomalies detected or AI analysis returned empty result")

        print(f"\n✅ File '{fname}' processing completed\n")

    # === CONSOLIDATION & EXCEL GENERATION ===
    print(f"\n📊 ===== CONSOLIDATING RESULTS =====\n")
    print(f"📈 Processed {len(files)} file(s) successfully")

    if all_anoms:
        print(f"🔗 Consolidating {len(all_anoms)} result set(s)...")
        anom_df = pd.concat(all_anoms, ignore_index=True)
        print(f"✅ Consolidation completed")
        print(f"   • Total anomalies: {len(anom_df)}")

        # Summary by subsidiary
        if len(anom_df) > 0:
            sub_summary = anom_df['Subsidiary'].value_counts()
            print(f"\n📋 Anomaly summary by subsidiary:")
            for sub, count in sub_summary.items():
                print(f"   • {sub}: {count} anomalies")

            status_summary = anom_df['Status'].value_counts()
            print(f"\n🔍 Analysis status summary:")
            for status, count in status_summary.items():
                print(f"   • {status}: {count}")
    else:
        print(f"⚠️  No anomalies detected across all files")
        anom_df = pd.DataFrame(columns=[
            "Subsidiary","Account","Period","Pct Change","Abs Change (VND)",
            "Trigger(s)","Suggested likely cause","Status","Notes"
        ])

    # === WRITE TO WORKSHEET ===
    print(f"\n📝 Writing results to Excel worksheet...")
    row_count = 0
    for r in dataframe_to_rows(anom_df, index=False, header=True):
        ws.append(r)
        row_count += 1
    print(f"✅ Written {row_count} rows to worksheet (including header)")

    # === VISUAL FORMATTING ===
    print(f"\n🎨 Applying visual formatting to Excel output...")
    apply_excel_formatting_ws(ws, anom_df, CONFIG)
    print(f"✅ Excel formatting applied successfully")

    # === RETURN BYTES ===
    print(f"\n💾 Generating final Excel file...")
    bio = io.BytesIO()
    wb.save(bio)
    final_size = len(bio.getvalue())
    print(f"✅ Excel file generated successfully")
    print(f"   • Output size: {final_size:,} bytes ({final_size/1024:.1f} KB)")

    print(f"\n📊 Debug Files Summary:")
    print(f"   • Debug files created: {len(debug_files)}")
    for debug_name, debug_bytes in debug_files:
        print(f"     - {debug_name}: {len(debug_bytes):,} bytes ({len(debug_bytes)/1024:.1f} KB)")

    print(f"\n🎉 ===== AI VARIANCE ANALYSIS COMPLETED =====\n")
    return bio.getvalue(), debug_files