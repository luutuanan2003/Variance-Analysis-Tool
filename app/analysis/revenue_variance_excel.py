# app/revenue_variance_excel.py
"""Excel formatting for revenue variance analysis."""

import pandas as pd

def _add_revenue_variance_analysis_to_sheet(ws, variance_analysis: dict):
    """Add the new comprehensive revenue variance analysis to an Excel worksheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(bold=True, size=12, color="2F5597")
    insight_font = Font(bold=True, size=11, color="D83B01")

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def format_vnd(amount):
        if isinstance(amount, (int, float)) and not pd.isna(amount):
            return f"{amount:,.0f} VND"
        return "N/A"

    def format_pct(pct):
        if isinstance(pct, (int, float)) and not pd.isna(pct):
            return f"{pct:+.1f}%"
        return "N/A"

    row = 1

    # Title
    ws[f"A{row}"] = "COMPREHENSIVE REVENUE VARIANCE ANALYSIS"
    ws[f"A{row}"].font = Font(bold=True, size=16, color="2F5597")
    row += 2

    # Executive Summary
    ws[f"A{row}"] = "EXECUTIVE SUMMARY"
    ws[f"A{row}"].font = section_font
    row += 1

    summary_data = [
        ["Subsidiary", variance_analysis.get('subsidiary', 'N/A')],
        ["File", variance_analysis.get('filename', 'N/A')],
        ["Months Analyzed", len(variance_analysis.get('months_analyzed', []))],
        ["Revenue Streams Identified", variance_analysis.get('analysis_summary', {}).get('total_revenue_streams', 0)],
        ["Variance Periods", variance_analysis.get('analysis_summary', {}).get('total_variance_periods', 0)],
        ["Accounts with Vendor Impact", variance_analysis.get('analysis_summary', {}).get('accounts_with_vendor_impact', 0)]
    ]

    for label, value in summary_data:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = Font(bold=True)
        row += 1
    row += 1

    # Key Insights
    if variance_analysis.get('key_insights'):
        ws[f"A{row}"] = "KEY INSIGHTS"
        ws[f"A{row}"].font = section_font
        row += 1

        for insight in variance_analysis['key_insights']:
            ws[f"A{row}"] = f"• {insight}"
            ws[f"A{row}"].font = insight_font
            row += 1
        row += 1

    # 1. Total Revenue Month-over-Month Changes
    if variance_analysis.get('total_revenue_analysis', {}).get('month_over_month_changes'):
        ws[f"A{row}"] = "1. TOTAL REVENUE MONTH-OVER-MONTH ANALYSIS"
        ws[f"A{row}"].font = section_font
        row += 1

        # Headers
        headers = ["Period From", "Period To", "Previous Revenue", "Current Revenue", "Absolute Change", "% Change", "Direction"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        changes = variance_analysis['total_revenue_analysis']['month_over_month_changes']
        for change in changes:
            ws[f"A{row}"] = change.get('period_from', '')
            ws[f"B{row}"] = change.get('period_to', '')
            ws[f"C{row}"] = format_vnd(change.get('previous_revenue', 0))
            ws[f"D{row}"] = format_vnd(change.get('current_revenue', 0))
            ws[f"E{row}"] = format_vnd(change.get('absolute_change', 0))
            ws[f"F{row}"] = format_pct(change.get('percentage_change', 0))
            ws[f"G{row}"] = change.get('change_direction', '')

            # Color code based on direction
            direction = change.get('change_direction', '').lower()
            if 'increase' in direction:
                fill_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            elif 'decrease' in direction:
                fill_color = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            else:
                fill_color = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill_color
                cell.border = thin_border
            row += 1
        row += 1

    # 2. Revenue Stream Analysis
    if variance_analysis.get('revenue_stream_analysis', {}).get('streams'):
        ws[f"A{row}"] = "2. REVENUE STREAM ANALYSIS (Contribution Code '01' Only)"
        ws[f"A{row}"].font = section_font
        row += 1

        streams = variance_analysis['revenue_stream_analysis']['streams']

        for stream_name, stream_data in streams.items():
            ws[f"A{row}"] = f"Stream: {stream_name}"
            ws[f"A{row}"].font = Font(bold=True, color="1F4E79")
            row += 1

            ws[f"A{row}"] = f"Total Entities: {stream_data.get('total_entities', 0)}"
            row += 1

            if stream_data.get('month_changes'):
                # Headers for month changes
                headers = ["Period From", "Period To", "Previous Value", "Current Value", "Change", "% Change"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    cell.border = thin_border
                row += 1

                for change in stream_data['month_changes']:
                    ws[f"A{row}"] = change.get('period_from', '')
                    ws[f"B{row}"] = change.get('period_to', '')
                    ws[f"C{row}"] = format_vnd(change.get('previous_value', 0))
                    ws[f"D{row}"] = format_vnd(change.get('current_value', 0))
                    ws[f"E{row}"] = format_vnd(change.get('absolute_change', 0))
                    ws[f"F{row}"] = format_pct(change.get('percentage_change', 0))

                    for col in range(1, 7):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border
                    row += 1
            row += 1

    # 3. Vendor/Customer Impact Analysis with Net Effect
    if variance_analysis.get('vendor_customer_impact', {}).get('detailed_analysis'):
        ws[f"A{row}"] = "3. VENDOR/CUSTOMER IMPACT ANALYSIS (Net Effect Breakdown)"
        ws[f"A{row}"].font = section_font
        row += 1

        detailed_analysis = variance_analysis['vendor_customer_impact']['detailed_analysis']

        for account_name, account_analysis in detailed_analysis.items():
            ws[f"A{row}"] = f"Account: {account_name}"
            ws[f"A{row}"].font = Font(bold=True, color="1F4E79")
            row += 1

            if account_analysis.get('period_impacts'):
                for period_impact in account_analysis['period_impacts']:
                    ws[f"A{row}"] = f"Period: {period_impact.get('period_from', '')} → {period_impact.get('period_to', '')}"
                    ws[f"A{row}"].font = Font(bold=True)
                    row += 1

                    # Net Effect Explanation
                    net_explanation = period_impact.get('net_effect_explanation', '')
                    ws[f"A{row}"] = f"Net Effect: {net_explanation}"
                    ws[f"A{row}"].font = Font(bold=True, color="D83B01")
                    row += 1

                    # Summary statistics
                    total_positive = period_impact.get('total_positive_change', 0)
                    total_negative = period_impact.get('total_negative_change', 0)
                    ws[f"A{row}"] = f"Total Increases: {format_vnd(total_positive)} | Total Decreases: {format_vnd(total_negative)} | Net: {format_vnd(total_positive + total_negative)}"
                    ws[f"A{row}"].font = Font(italic=True)
                    row += 1

                    ws[f"A{row}"] = f"Entities with Significant Changes: {period_impact.get('entities_with_significant_change', 0)}"
                    row += 1
                    row += 1

                    # POSITIVE CONTRIBUTORS Section
                    if period_impact.get('positive_contributors'):
                        ws[f"A{row}"] = "POSITIVE CONTRIBUTORS (Increases)"
                        ws[f"A{row}"].font = Font(bold=True, color="0F7B0F")
                        row += 1

                        # Headers for positive contributions
                        headers = ["Entity", "Previous Value", "Current Value", "Increase (+)", "% Change", "% of Total Change"]
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row=row, column=col, value=header)
                            cell.font = header_font
                            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                            cell.border = thin_border
                        row += 1

                        for entity in period_impact['positive_contributors']:
                            ws[f"A{row}"] = entity.get('entity', '')
                            ws[f"B{row}"] = format_vnd(entity.get('previous_value', 0))
                            ws[f"C{row}"] = format_vnd(entity.get('current_value', 0))
                            ws[f"D{row}"] = f"+{entity.get('absolute_change', 0):,.0f} VND"
                            ws[f"E{row}"] = format_pct(entity.get('percentage_change', 0))
                            ws[f"F{row}"] = format_pct(entity.get('contribution_to_period_change', 0))

                            # Green fill for positive contributors
                            fill_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                            for col in range(1, 7):
                                cell = ws.cell(row=row, column=col)
                                cell.fill = fill_color
                                cell.border = thin_border
                            row += 1
                        row += 1

                    # NEGATIVE CONTRIBUTORS Section
                    if period_impact.get('negative_contributors'):
                        ws[f"A{row}"] = "NEGATIVE CONTRIBUTORS (Decreases)"
                        ws[f"A{row}"].font = Font(bold=True, color="C5504B")
                        row += 1

                        # Headers for negative contributions
                        headers = ["Entity", "Previous Value", "Current Value", "Decrease (-)", "% Change", "% of Total Change"]
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row=row, column=col, value=header)
                            cell.font = header_font
                            cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
                            cell.border = thin_border
                        row += 1

                        for entity in period_impact['negative_contributors']:
                            ws[f"A{row}"] = entity.get('entity', '')
                            ws[f"B{row}"] = format_vnd(entity.get('previous_value', 0))
                            ws[f"C{row}"] = format_vnd(entity.get('current_value', 0))
                            ws[f"D{row}"] = f"{entity.get('absolute_change', 0):,.0f} VND"  # Already negative
                            ws[f"E{row}"] = format_pct(entity.get('percentage_change', 0))
                            ws[f"F{row}"] = format_pct(entity.get('contribution_to_period_change', 0))

                            # Red fill for negative contributors
                            fill_color = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                            for col in range(1, 7):
                                cell = ws.cell(row=row, column=col)
                                cell.fill = fill_color
                                cell.border = thin_border
                            row += 1
                        row += 1

                    # NET EFFECT VALIDATION
                    calculated_net = period_impact.get('net_effect_calculated', 0)
                    actual_total = period_impact.get('total_change', 0)
                    ws[f"A{row}"] = f"Validation: Calculated Net = {format_vnd(calculated_net)} | Actual Total = {format_vnd(actual_total)}"
                    if abs(calculated_net - actual_total) < 1:  # Allow for rounding
                        ws[f"A{row}"].font = Font(color="0F7B0F")
                    else:
                        ws[f"A{row}"].font = Font(color="C5504B")
                    row += 1
                    row += 1
            row += 1

    # Configuration Used
    if variance_analysis.get('configuration_used'):
        ws[f"A{row}"] = "CONFIGURATION USED"
        ws[f"A{row}"].font = section_font
        row += 1

        config = variance_analysis['configuration_used']
        config_data = [
            ["Months Analyzed", config.get('months_analyzed', 'N/A')],
            ["Revenue Entity Threshold (VND)", format_vnd(config.get('revenue_entity_threshold_vnd', 0))],
            ["Excluded Base Account", config.get('excluded_base_account', 'N/A')]
        ]

        for label, value in config_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
        row += 1