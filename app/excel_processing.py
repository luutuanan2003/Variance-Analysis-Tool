# app/excel_processing.py
"""Excel processing and formatting functions."""

from __future__ import annotations

import io
import re
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from .data_utils import (
    detect_header_row, normalize_financial_col, promote_row8,
    fill_down_assign, coerce_numeric, aggregate_totals, DEFAULT_CONFIG
)

# ---------------------------
# Helpers for grouped sheets
# ---------------------------

def _short_title(name: str, CONFIG: dict = DEFAULT_CONFIG) -> str:
    # Excel sheet name rules
    name = re.sub(r'[:\\/?*\[\]]', '-', name).strip()
    return name[:CONFIG.get("max_sheet_name_length", 31)]

def _format_pct(pct):
    return "N/A" if pct is None else f"{pct:+.1f}%"

def _write_account_block(
    ws: Worksheet,
    row: int,
    account_name: str,
    account_data: dict,
    *,
    section_font,
    header_font,
    header_fill,
    thin_border,
    format_vnd,
    cost_mode: bool,  # True for 632/641/642 (up = bad/red), False for 511 (up = good/green)
) -> int:
    # Account header
    ws[f"A{row}"] = f"Account: {account_name}"
    ws[f"A{row}"].font = section_font
    row += 1

    biggest = account_data.get("biggest_change")
    if biggest:
        ws[f"A{row}"] = "Biggest Change"
        ws[f"B{row}"] = f"{biggest.get('from','')} → {biggest.get('to','')}"
        ws[f"C{row}"] = format_vnd(biggest.get('change', 0))
        ws[f"D{row}"] = _format_pct(biggest.get('pct_change'))
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
        row += 1

    # MoM table
    headers = ["Period", "Prev (VND)", "Curr (VND)", "Δ (VND)", "Δ (%)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1

    for ch in account_data.get("changes", []):
        ws[f"A{row}"] = f"{ch.get('from','')} → {ch.get('to','')}"
        ws[f"B{row}"] = format_vnd(ch.get('prev_val', 0))
        ws[f"C{row}"] = format_vnd(ch.get('curr_val', 0))
        ws[f"D{row}"] = format_vnd(ch.get('change', 0))
        ws[f"E{row}"] = _format_pct(ch.get('pct_change'))

        delta = ch.get('change', 0)
        # polarity: revenue up = good; cost up = bad
        if delta != 0:
            good = (delta > 0) if not cost_mode else (delta < 0)
            fill = PatternFill(
                start_color=("E8F5E8" if good else "FFEBEE"),
                end_color=("E8F5E8" if good else "FFEBEE"),
                fill_type="solid",
            )
        else:
            fill = None

        for col in range(1, 6):
            c = ws.cell(row=row, column=col)
            if fill:
                c.fill = fill
            c.border = thin_border
        row += 1

    # Top entity impacts
    impacts = account_data.get("customer_impacts") or account_data.get("entity_impacts") or []
    if impacts:
        row += 1
        ws[f"A{row}"] = "Top Entity Impacts:"
        ws[f"A{row}"].font = section_font
        row += 1

        heads = ["Entity", "Δ (VND)", "Δ (%)", "Prev", "Curr"]
        for col, h in enumerate(heads, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        for imp in impacts:
            ws[f"A{row}"] = imp.get("entity", "")
            ws[f"B{row}"] = format_vnd(imp.get("change", 0))
            ws[f"C{row}"] = _format_pct(imp.get("pct_change"))
            ws[f"D{row}"] = format_vnd(imp.get("prev_val", 0))
            ws[f"E{row}"] = format_vnd(imp.get("curr_val", 0))

            delta = imp.get('change', 0)
            if delta != 0:
                good = (delta > 0) if not cost_mode else (delta < 0)
                fill = PatternFill(
                    start_color=("E8F5E8" if good else "FFEBEE"),
                    end_color=("E8F5E8" if good else "FFEBEE"),
                    fill_type="solid",
                )
            else:
                fill = None

            for col in range(1, 6):
                c = ws.cell(row=row, column=col)
                if fill:
                    c.fill = fill
                c.border = thin_border
            row += 1

    row += 2
    return row

def _autofit_ws(ws: Worksheet):
    for column in ws.columns:
        maxlen = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                maxlen = max(maxlen, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(maxlen + 2, 50)

def create_group_sheet(
    wb,
    title: str,
    accounts: dict,
    *,
    section_font,
    header_font,
    header_fill,
    thin_border,
    format_vnd,
    cost_mode: bool,
):
    title = _short_title(title)
    ws = wb.create_sheet(title)
    row = 1
    ws[f"A{row}"] = title.upper()
    ws[f"A{row}"].font = section_font
    row += 2

    # sort by |biggest change| desc
    sorted_items = []
    for acct, data in accounts.items():
        big = data.get("biggest_change") or {}
        sorted_items.append((acct, abs(big.get("change", 0)), data))
    sorted_items.sort(key=lambda t: t[1], reverse=True)

    for acct, _, data in sorted_items:
        row = _write_account_block(
            ws, row, acct, data,
            section_font=section_font,
            header_font=header_font,
            header_fill=header_fill,
            thin_border=thin_border,
            format_vnd=format_vnd,
            cost_mode=cost_mode,
        )

    _autofit_ws(ws)

# --------------------------------
# Sheet load/normalize utilities
# --------------------------------

def process_financial_tab_from_bytes(
    xl_bytes: bytes,
    sheet_name: str,
    mode: str,
    subsidiary: str,
) -> tuple[pd.DataFrame, list[str]]:
    """Load and clean one sheet ('BS Breakdown' or 'PL Breakdown') from in-memory bytes."""
    header_row = detect_header_row(xl_bytes, sheet_name)
    df = pd.read_excel(io.BytesIO(xl_bytes), sheet_name=sheet_name, header=header_row, dtype=str)
    df = normalize_financial_col(df)
    df, month_cols = promote_row8(df, mode, subsidiary)
    df = fill_down_assign(df)
    df = coerce_numeric(df, month_cols)
    keep_cols = ["Account Code","Account Name","RowHadOwnCode","IsTotal"] + [c for c in month_cols if c in df.columns]
    df = df[keep_cols]
    totals = aggregate_totals(df, month_cols)
    return totals, month_cols

def extract_subsidiary_name_from_bytes(xl_bytes: bytes, fallback_filename: str) -> str:
    """Try to find a name on A2 of BS/PL sheets like 'Subsidiary: XYZ'. Fallback to filename stem."""
    try:
        wb = load_workbook(io.BytesIO(xl_bytes), read_only=True, data_only=True)
        for sheet_name in ["BS Breakdown", "PL Breakdown"]:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                cell_value = sheet["A2"].value
                if isinstance(cell_value, str) and ":" in cell_value:
                    wb.close()
                    return cell_value.split(":")[-1].strip()
        wb.close()
    except Exception:
        pass
    # fallback: filename before first underscore or dot
    stem = fallback_filename.rsplit("/", 1)[-1]
    stem = stem.split("\\")[-1]
    stem = stem.split(".")[0]
    return stem.split("_")[0] if "_" in stem else stem

# -----------------------------------------------------------------------------
# Excel formatting (IN-MEMORY, works on a worksheet not a saved file)
# -----------------------------------------------------------------------------

def apply_excel_formatting_ws(ws, anomaly_df: pd.DataFrame, CONFIG: dict) -> None:
    """Apply simple conditional fills directly on the 'Anomalies Summary' worksheet."""
    try:
        critical_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        warning_fill  = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        header_fill   = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        ai_fill       = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")

        # Header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)

        # Find indexes
        headers = [c.value for c in ws[1]]
        try:
            abs_idx = headers.index("Abs Change (VND)") + 1
            trig_idx = headers.index("Trigger(s)") + 1
            status_idx = headers.index("Status") + 1
        except ValueError:
            return

        # Rows
        for row_idx in range(2, ws.max_row + 1):
            try:
                abs_change = ws.cell(row=row_idx, column=abs_idx).value or 0
                trigger = str(ws.cell(row=row_idx, column=trig_idx).value or "")
                status = str(ws.cell(row=row_idx, column=status_idx).value or "")

                fill = None
                if "AI Analysis" in status:
                    fill = ai_fill
                elif abs_change >= CONFIG.get("materiality_vnd", 1000000000) * 5:
                    fill = critical_fill
                elif "Correlation break" in trigger or abs_change >= CONFIG.get("materiality_vnd", 1000000000) * 2:
                    fill = warning_fill

                if fill:
                    for col_idx in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill
            except Exception:
                continue
    except Exception:
        # Formatting should never break
        pass

# -------------------------------------------
# Main: write summary sheet + grouped sheets
# -------------------------------------------

def _add_revenue_analysis_to_sheet(ws, revenue_analysis: dict):
    """Add revenue analysis data to an Excel worksheet in a structured format."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(bold=True, size=12, color="2F5597")
    currency_font = Font(name="Arial", size=10)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def format_vnd(amount):
        if isinstance(amount, (int, float)) and not pd.isna(amount):
            return f"{amount:,.0f} VND"
        return "N/A"

    row = 1

    # Title
    ws[f"A{row}"] = "COMPREHENSIVE REVENUE ANALYSIS"
    ws[f"A{row}"].font = Font(bold=True, size=16, color="2F5597")
    row += 2

    # Executive Summary
    if revenue_analysis.get('summary'):
        summary = revenue_analysis['summary']
        ws[f"A{row}"] = "EXECUTIVE SUMMARY"
        ws[f"A{row}"].font = section_font
        row += 1

        summary_data = [
            ["Subsidiary", revenue_analysis.get('subsidiary', 'N/A')],
            ["Months Analyzed", len(revenue_analysis.get('months_analyzed', []))],
            ["Revenue Accounts", summary.get('total_accounts', 0)],
            ["Latest Total Revenue", format_vnd(summary.get('total_revenue_latest', 0))],
            ["Latest Gross Margin %", f"{summary.get('gross_margin_latest', 0):.1f}%" if summary.get('gross_margin_latest') else 'N/A'],
            ["SG&A 641* Accounts", summary.get('total_sga_641_accounts', 0)],
            ["SG&A 642* Accounts", summary.get('total_sga_642_accounts', 0)],
            ["Latest Total SG&A", format_vnd(summary.get('total_sga_latest', 0))],
            ["Latest SG&A Ratio %", f"{summary.get('sga_ratio_latest', 0):.1f}%" if summary.get('sga_ratio_latest') else 'N/A']
        ]

        for label, value in summary_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
        row += 1

    # Risk Assessment
    if revenue_analysis.get('risk_assessment'):
        ws[f"A{row}"] = "RISK ASSESSMENT"
        ws[f"A{row}"].font = section_font
        row += 1

        # Headers
        headers = ["Period", "Risk Level", "Type", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        for risk in revenue_analysis['risk_assessment']:
            ws[f"A{row}"] = risk.get('period', '')
            ws[f"B{row}"] = risk.get('risk_level', '')
            ws[f"C{row}"] = risk.get('type', '')
            ws[f"D{row}"] = risk.get('description', '')

            # Color code risk levels
            risk_level = risk.get('risk_level', '').lower()
            if risk_level == 'high':
                fill_color = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            elif risk_level == 'medium':
                fill_color = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            else:
                fill_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")

            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill_color
                cell.border = thin_border
            row += 1
        row += 1

    # Total Revenue Trend
    if revenue_analysis.get('total_revenue_analysis', {}).get('changes'):
        ws[f"A{row}"] = "TOTAL REVENUE TREND (511*)"
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["Period", "Previous Value", "Current Value", "Change (VND)", "Change (%)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        for change in revenue_analysis['total_revenue_analysis']['changes']:
            period = f"{change.get('from', '')} → {change.get('to', '')}"
            ws[f"A{row}"] = period
            ws[f"B{row}"] = format_vnd(change.get('prev_value', 0))
            ws[f"C{row}"] = format_vnd(change.get('curr_value', 0))
            ws[f"D{row}"] = format_vnd(change.get('change', 0))
            ws[f"E{row}"] = f"{change.get('pct_change', 0):+.1f}%"

            change_val = change.get('change', 0)
            if change_val > 0:
                fill_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            elif change_val < 0:
                fill_color = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            else:
                fill_color = None

            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                if fill_color:
                    cell.fill = fill_color
                cell.border = thin_border
            row += 1
        row += 1

        # Total COGS Trend (632*)
    if revenue_analysis.get('total_632_trend', {}).get('changes'):
        ws[f"A{row}"] = "TOTAL COGS TREND (632*)"
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["Period", "Previous Value", "Current Value", "Change (VND)", "Change (%)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        for change in revenue_analysis['total_632_trend']['changes']:
            ws[f"A{row}"] = f"{change.get('from', '')} → {change.get('to', '')}"
            ws[f"B{row}"] = format_vnd(change.get('prev_value', 0))
            ws[f"C{row}"] = format_vnd(change.get('curr_value', 0))
            ws[f"D{row}"] = format_vnd(change.get('change', 0))
            ws[f"E{row}"] = f"{change.get('pct_change', 0):+.1f}%"

            delta = change.get('change', 0)
            if delta > 0:
                fill_color = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")  # Cost up = bad
            elif delta < 0:
                fill_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")  # Cost down = good
            else:
                fill_color = None

            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                if fill_color:
                    cell.fill = fill_color
                cell.border = thin_border
            row += 1
        row += 1

    # Total SG&A 641 Trend
    if revenue_analysis.get('total_641_trend', {}).get('changes'):
        ws[f"A{row}"] = "TOTAL SG&A TREND (641*)"
        ws[f"A{row}"].font = section_font
        row += 1
        # Reuse the same header and rendering pattern as above
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1
        for change in revenue_analysis['total_641_trend']['changes']:
            ws[f"A{row}"] = f"{change.get('from', '')} → {change.get('to', '')}"
            ws[f"B{row}"] = format_vnd(change.get('prev_value', 0))
            ws[f"C{row}"] = format_vnd(change.get('curr_value', 0))
            ws[f"D{row}"] = format_vnd(change.get('change', 0))
            ws[f"E{row}"] = f"{change.get('pct_change', 0):+.1f}%"
            delta = change.get('change', 0)
            fill_color = PatternFill(start_color="FFEBEE" if delta > 0 else "E8F5E8", end_color="FFEBEE" if delta > 0 else "E8F5E8", fill_type="solid") if delta != 0 else None
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                if fill_color:
                    cell.fill = fill_color
                cell.border = thin_border
            row += 1
        row += 1

    # Total SG&A 642 Trend
    if revenue_analysis.get('total_642_trend', {}).get('changes'):
        ws[f"A{row}"] = "TOTAL SG&A TREND (642*)"
        ws[f"A{row}"].font = section_font
        row += 1
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1
        for change in revenue_analysis['total_642_trend']['changes']:
            ws[f"A{row}"] = f"{change.get('from', '')} → {change.get('to', '')}"
            ws[f"B{row}"] = format_vnd(change.get('prev_value', 0))
            ws[f"C{row}"] = format_vnd(change.get('curr_value', 0))
            ws[f"D{row}"] = format_vnd(change.get('change', 0))
            ws[f"E{row}"] = f"{change.get('pct_change', 0):+.1f}%"
            delta = change.get('change', 0)
            fill_color = PatternFill(start_color="FFEBEE" if delta > 0 else "E8F5E8", end_color="FFEBEE" if delta > 0 else "E8F5E8", fill_type="solid") if delta != 0 else None
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                if fill_color:
                    cell.fill = fill_color
                cell.border = thin_border
            row += 1
        row += 1


    # Gross Margin Analysis
    if revenue_analysis.get('gross_margin_analysis', {}).get('trend'):
        ws[f"A{row}"] = "GROSS MARGIN ANALYSIS"
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["Month", "Revenue", "Cost", "Gross Margin %", "Change from Previous"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        prev_margin = None
        for margin_data in revenue_analysis['gross_margin_analysis']['trend']:
            ws[f"A{row}"] = margin_data.get('month', '')
            ws[f"B{row}"] = format_vnd(margin_data.get('revenue', 0))
            ws[f"C{row}"] = format_vnd(margin_data.get('cost', 0))
            ws[f"D{row}"] = f"{margin_data.get('gross_margin_pct', 0):.1f}%"

            current_margin = margin_data.get('gross_margin_pct', 0)
            if prev_margin is not None:
                change = current_margin - prev_margin
                ws[f"E{row}"] = f"{change:+.1f}pp"
                if change > 0:
                    ws.cell(row=row, column=5).fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                elif change < 0:
                    ws.cell(row=row, column=5).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            else:
                ws[f"E{row}"] = "N/A"

            prev_margin = current_margin

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = thin_border
            row += 1
        row += 1

    # Utility Analysis (if available)
    if revenue_analysis.get('utility_analysis'):
        ws[f"A{row}"] = "UTILITY REVENUE VS COST ANALYSIS"
        ws[f"A{row}"].font = section_font
        row += 1

        if revenue_analysis['utility_analysis'].get('available') and revenue_analysis['utility_analysis'].get('margins'):
            headers = ["Month", "Utility Revenue", "Utility Cost", "Margin (VND)", "Margin %"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            row += 1

            for margin in revenue_analysis['utility_analysis']['margins']:
                ws[f"A{row}"] = margin.get('month', '')
                ws[f"B{row}"] = format_vnd(margin.get('revenue', 0))
                ws[f"C{row}"] = format_vnd(margin.get('cost', 0))
                ws[f"D{row}"] = format_vnd(margin.get('revenue', 0) - margin.get('cost', 0))
                ws[f"E{row}"] = f"{margin.get('margin_pct', 0):.1f}%"

                margin_pct = margin.get('margin_pct', 0)
                if margin_pct >= 0:
                    fill_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                else:
                    fill_color = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill_color
                    cell.border = thin_border
                row += 1
        else:
            ws[f"A{row}"] = "Utility accounts not found in the data."
            ws[f"A{row}"].font = Font(italic=True, color="666666")
            row += 1
        row += 1

    # Combined SG&A Analysis (641 + 642)
    if revenue_analysis.get('combined_sga_analysis', {}).get('ratio_trend'):
        ws[f"A{row}"] = "COMBINED SG&A ANALYSIS (641* + 642*)"
        ws[f"A{row}"].font = section_font
        row += 1

        combined_analysis = revenue_analysis['combined_sga_analysis']
        ws[f"A{row}"] = f"Total 641* Accounts: {combined_analysis.get('total_641_accounts', 0)}"
        row += 1
        ws[f"A{row}"] = f"Total 642* Accounts: {combined_analysis.get('total_642_accounts', 0)}"
        row += 2

        ws[f"A{row}"] = "SG&A Ratio Trend (% of Revenue)"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        headers = ["Month", "Revenue", "641* Total", "642* Total", "Total SG&A", "SG&A Ratio %", "Change"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        prev_ratio = None
        for sga_data in combined_analysis['ratio_trend']:
            ws[f"A{row}"] = sga_data.get('month', '')
            ws[f"B{row}"] = format_vnd(sga_data.get('revenue', 0))
            ws[f"C{row}"] = format_vnd(sga_data.get('sga_641_total', 0))
            ws[f"D{row}"] = format_vnd(sga_data.get('sga_642_total', 0))
            ws[f"E{row}"] = format_vnd(sga_data.get('total_sga', 0))
            ws[f"F{row}"] = f"{sga_data.get('sga_ratio_pct', 0):.1f}%"

            current_ratio = sga_data.get('sga_ratio_pct', 0)
            if prev_ratio is not None:
                change = current_ratio - prev_ratio
                ws[f"G{row}"] = f"{change:+.1f}pp"
                if change > 2:
                    ws.cell(row=row, column=7).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                elif change < -2:
                    ws.cell(row=row, column=7).fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            else:
                ws[f"G{row}"] = "N/A"

            prev_ratio = current_ratio

            for col in range(1, 8):
                ws.cell(row=row, column=col).border = thin_border
            row += 1
        row += 1

    # Auto-fit main sheet
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # === GROUPED DETAIL SHEETS ===
    wb = ws.parent  # workbook

    if revenue_analysis.get('revenue_by_account'):
        create_group_sheet(
            wb, "511 Accounts",
            revenue_analysis['revenue_by_account'],
            section_font=section_font,
            header_font=header_font,
            header_fill=header_fill,
            thin_border=thin_border,
            format_vnd=format_vnd,
            cost_mode=False,  # revenue polarity
        )

    if revenue_analysis.get('cogs_632_analysis'):
        create_group_sheet(
            wb, "632 Accounts",
            revenue_analysis['cogs_632_analysis'],
            section_font=section_font,
            header_font=header_font,
            header_fill=header_fill,
            thin_border=thin_border,
            format_vnd=format_vnd,
            cost_mode=True,
        )

    if revenue_analysis.get('sga_641_analysis'):
        create_group_sheet(
            wb, "641 Accounts",
            revenue_analysis['sga_641_analysis'],
            section_font=section_font,
            header_font=header_font,
            header_fill=header_fill,
            thin_border=thin_border,
            format_vnd=format_vnd,
            cost_mode=True,
        )

    if revenue_analysis.get('sga_642_analysis'):
        create_group_sheet(
            wb, "642 Accounts",
            revenue_analysis['sga_642_analysis'],
            section_font=section_font,
            header_font=header_font,
            header_fill=header_fill,
            thin_border=thin_border,
            format_vnd=format_vnd,
            cost_mode=True,
        )
