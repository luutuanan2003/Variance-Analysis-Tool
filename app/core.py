from __future__ import annotations

import io
import re
import warnings
from typing import List, Tuple, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings("ignore")  # Avoid noisy pandas dtype warnings in logs

# Import AI analyzer for AI mode
try:
    from .llm_analyzer import LLMFinancialAnalyzer
    AI_AVAILABLE = True
except ImportError:
    AI_AVAILABLE = False
    print("⚠️  AI analyzer not available - AI mode will be disabled")

# === Accounting thresholds (centralized) ===
ACCT_THRESH = {
    # the % change in gross margin compared to another period
    # (last month, last year, or budget).
    "gross_margin_pct_delta": 0.01,    # 1% point change m/m
    # the percentage change in depreciation compared to a prior period
    # (last month, last year, or budget).
    "depr_pct_delta": 0.10,            # 10% change m/m for 217*, 632*, 214
    # how much the cost-to-revenue ratio has increased or decreased compared
    # to a previous period.
    "cogs_ratio_delta": 0.02,          # 2% points drift vs hist
    # how much the % of revenue spent on overhead (SGA) has increased or decreased
    # compared to history.
    "sga_pct_of_rev_delta": 0.10,      # +10% vs hist % of revenue
    # the % change (or volatility) in financial income/expenses compared to a prior period,
    # used to flag unusual financial fluctuations.
    "fin_swing_pct": 0.50,             # >50% swings
    # the % difference between depreciation recorded in the Balance Sheet vs.
    # the depreciation shown in the P&L.
    "bs_pl_dep_diff_pct": 0.05,        # 5% mismatch between 214/217 Δ and 632 dep expense
}

# === Account prefix helpers (VN CoA style) ===
def _is_511(name: str) -> bool:
    s = str(name).replace(" ", "").lower()
    return s.startswith("511")

def _is_632(name: str) -> bool:
    s = str(name).replace(" ", "").lower()
    return s.startswith("632")

def _is_641(name: str) -> bool:
    s = str(name).replace(" ", "").lower()
    return s.startswith("641")

def _is_642(name: str) -> bool:
    s = str(name).replace(" ", "").lower()
    return s.startswith("642")

def _is_635(name: str) -> bool:
    s = str(name).replace(" ", "").lower()
    return s.startswith("635")

def _is_515(name: str) -> bool:
    s = str(name).replace(" ", "").lower()
    return s.startswith("515")

def _is_217(name: str) -> bool:
    s = str(name).replace(" ", "").lower()
    return s.startswith("217")

def _is_214(name: str) -> bool:
    s = str(name).replace(" ", "").lower()
    return s.startswith("214")

# === Safe % change ===
def _pct_change(a: pd.Series) -> pd.Series:
    b = pd.to_numeric(a, errors="coerce")
    return b.pct_change().replace([np.inf, -np.inf], np.nan)

def _series_hist_pct_of_rev(series: pd.Series, rev: pd.Series) -> tuple[float, float]:
    """Return (hist_mean, hist_std) for (series / rev)."""
    x = pd.to_numeric(series, errors="coerce")
    r = pd.to_numeric(rev, errors="coerce")
    ratio = x / r.replace({0: np.nan})
    return ratio.mean(skipna=True), ratio.std(skipna=True)

def _months(df: pd.DataFrame) -> list[str]:
    # Assumes your pipeline already normalized to monthly columns like 'Jan 2025' ... 'Dec 2025'
    return [c for c in df.columns if isinstance(c, str) and c.strip().lower().endswith(("2024", "2025", "2026", "2027", "2028", "2029", "2030"))]

# === Output record helper ===
def _anom_record(rule: str, entity: str, account: str, month: str, value, detail: dict) -> dict:
    rec = {
        "Rule": rule,
        "Entity": entity,
        "Account": account,
        "Month": month,
        "Value": value,
    }
    rec.update(detail or {})
    return rec

# -----------------------------------------------------------------------------
# Defaults & constants (shared between Python and AI modes)
# -----------------------------------------------------------------------------

DEFAULT_CONFIG: dict = {
    # if an error or difference is bigger than this, it matters; if smaller, we can ignore it.
    "materiality_vnd": 1_000_000_000,      # absolute VND change threshold
    # the % cut-off used to decide if recurring revenue or costs are large enough (vs. total) to count as meaningful.
    "recurring_pct_threshold": 0.05,       # 5% for recurring P/L accounts
    # the % of revenue spent on operating expenses beyond which costs are considered too high.
    "revenue_opex_pct_threshold": 0.10,    # 10% for revenue/opex accounts
    # the % cut-off used to decide if a balance sheet account change is large enough (vs. total) to count as meaningful.
    "bs_pct_threshold": 0.05,              # 5% for balance sheet
    # list of account code prefixes that indicate recurring revenue or costs.
    "recurring_code_prefixes": ["6321", "635", "515"],
    # the minimum number of periods (months) with data required to perform trend analysis.
    "min_trend_periods": 3,
    # the % drop in gross margin (absolute points) that triggers an anomaly.
    "gm_drop_threshold_pct": 0.01,         # 1% absolute drop triggers (e.g., 0.01 = 1pp)
    # list of account code prefixes for which depreciation should be analyzed using only % change rules.
    "dep_pct_only_prefixes": ["217", "632"],  # treat these as %-only rules
    # list of keywords to identify a "customer" column in the P&L data.
    "customer_column_hints": ["customer", "khách", "khach", "client", "buyer", "entity", "company", "subsidiary", "parent company", "bwid", "vc1", "vc2", "vc3", "logistics"],  # for 511* drilldown
    # AI-specific configuration
    "use_llm_analysis": False,              # Whether to use AI analysis
    "llm_model": "gpt-4o",                  # LLM model for AI analysis
}

MONTHS = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]

# Patterns to recognize header row content that carries a period like "As of Feb-2024" etc.
BS_PAT = re.compile(r'^\s*as\s*of\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[\.\-\s]*(\d{2,4})\s*$', re.I)
PL_PAT = re.compile(r'^\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[\.\-\s]*(\d{2,4})\s*$', re.I)

# -----------------------------------------------------------------------------
# Helpers: find optional "customer" column
# -----------------------------------------------------------------------------
def find_customer_column(df: pd.DataFrame, CONFIG: dict) -> Optional[str]:
    """Try to find a customer dimension column in a cleaned PL frame.
    Heuristic: any column whose name contains one of the configured hints.
    """
    if df is None or df.empty:
        return None
    cols = [str(c).strip() for c in df.columns]
    hints = [h.lower() for h in CONFIG.get("customer_column_hints", [])]
    for c in cols:
        lc = c.lower()
        if any(h in lc for h in hints):
            return c
    return None

# -----------------------------------------------------------------------------
# Helpers: period parsing / ordering
# -----------------------------------------------------------------------------
def normalize_period_label(label: object) -> str:
    """Turn many month-year formats into 'Mon YYYY'."""
    if label is None:
        return ""
    s = str(label).strip()
    if s == "":
        return ""
    try:
        s_clean = re.sub(r'^\s*(as\s*of|tinh\s*den|tính\s*đến|den\s*ngay|đến\s*ngày)\s*', '', s, flags=re.I)

        m = re.search(r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[^\w]?[\s\-\.]*([12]\d{3}|\d{2})\b',
                      s_clean, flags=re.I)
        if m:
            mon, yr = m.group(1), m.group(2)
            yr = int(yr)
            yr = yr + 2000 if yr < 100 else yr
            return f"{mon.title()} {yr}"

        m = re.search(r'\b(1[0-2]|0?[1-9])[./\-](\d{4})\b', s_clean)
        if m:
            mon = int(m.group(1)); yr = int(m.group(2))
            return f"{MONTHS[mon-1].title()} {yr}"

        m = re.search(r'\b(\d{4})[./\-](1[0-2]|0?[1-9])\b', s_clean)
        if m:
            yr = int(m.group(1)); mon = int(m.group(2))
            return f"{MONTHS[mon-1].title()} {yr}"

        m_year = re.search(r'(20\d{2}|19\d{2})', s_clean)
        m_mon  = re.search(r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\b', s_clean, flags=re.I)
        if m_year and m_mon:
            yr  = int(m_year.group(1))
            mon = m_mon.group(0)
            return f"{mon.title()} {yr}"
    except Exception:
        pass
    return s

def month_key(label: object) -> tuple[int, int]:
    """Return (year, month) for sorting. Unknown -> (9999, 99)."""
    n = normalize_period_label(label)
    m = re.search(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+(\d{4})', n, re.I)
    if not m:
        return (9999, 99)
    y = int(m.group(2))
    mi = MONTHS.index(m.group(1).lower()) + 1
    return (y, mi)

# -----------------------------------------------------------------------------
# Excel reading / header detection / cleaning (IN-MEMORY)
# -----------------------------------------------------------------------------

def detect_header_row(xl_bytes: bytes, sheet: str) -> int:
    """Heuristically find the header row by scanning first ~40 rows for 'Financial row'."""
    try:
        probe = pd.read_excel(io.BytesIO(xl_bytes), sheet_name=sheet, header=None, nrows=40)
        for i in range(len(probe)):
            row_values = probe.iloc[i].astype(str).str.strip().str.lower()
            if any("financial row" in v for v in row_values):
                return i
    except Exception:
        pass
    return 0  # fallback to first row

def normalize_financial_col(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure the main descriptor column is exactly 'Financial row'."""
    for c in df.columns:
        if str(c).strip().lower() == "financial row":
            return df.rename(columns={c: "Financial row"})
    # Otherwise assume first column is the descriptor
    return df.rename(columns={df.columns[0]: "Financial row"})

def promote_row8(df: pd.DataFrame, mode: str, sub: str) -> tuple[pd.DataFrame, list[str]]:
    """Use the first data row as headers when period info is there; normalize month columns."""
    if len(df) < 1:
        return df, []
    row0 = df.iloc[0]
    new_cols: list[str] = []
    for c in df.columns:
        v = str(row0.get(c, "")).strip()
        if BS_PAT.match(v) or PL_PAT.match(v):
            new_cols.append(normalize_period_label(v))
        else:
            new_cols.append(str(c))
    df = df.copy()
    df.columns = new_cols
    df = df.iloc[1:].reset_index(drop=True)

    month_cols: list[str] = []
    for c in df.columns:
        normalized = normalize_period_label(c)
        if re.match(r'^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+\d{4}$', normalized, re.I):
            month_cols.append(c)
    month_cols = sorted(month_cols, key=month_key)
    return df, month_cols

def fill_down_assign(df: pd.DataFrame) -> pd.DataFrame:
    """Extract account code / name from descriptor and forward-fill."""
    ser = df["Financial row"].astype(str)

    code_extract = ser.str.extract(r'(\d{4,})', expand=False)
    name_extract = ser.str.replace(r'.*?(\d{4,})\s*[-:]*\s*', '', regex=True).str.strip()

    row_has_code     = code_extract.notna()
    is_total_word    = ser.str.strip().str.lower().str.startswith(("total","subtotal","cộng","tong","tổng"))
    is_total_with_code = is_total_word & row_has_code
    is_section       = ser.str.match(r'^\s*([IVX]+\.|[A-Z]\.)\s')
    is_empty         = ser.str.strip().eq("")

    df["Account Code"] = code_extract.ffill()
    df["Account Name"] = name_extract.where(row_has_code).ffill()
    df["RowHadOwnCode"] = row_has_code
    df["IsTotal"] = is_total_with_code

    keep_mask = ~(is_section | is_empty)
    df = df[keep_mask & df["Account Code"].notna()].copy()
    return df

def coerce_numeric(df: pd.DataFrame, month_cols: list[str]) -> pd.DataFrame:
    """Coerce month columns to numeric."""
    out = df.copy()
    for c in month_cols:
        if c in out.columns:
            series = out[c].astype(str)
            series = (
                series
                .str.replace("\u00a0","", regex=False)    # nbsp
                .str.replace(",","", regex=False)
                .str.replace(r"\((.*)\)", r"-\1", regex=True)  # (100) -> -100
                .str.replace(r"[^0-9\.\-]", "", regex=True)
            )
            out[c] = pd.to_numeric(series, errors="coerce").fillna(0.0)
    return out

def aggregate_totals(df: pd.DataFrame, month_cols: list[str]) -> pd.DataFrame:
    """Aggregate to a single row per account code; prefer explicit 'total' lines when present."""
    if df.empty:
        return pd.DataFrame(columns=["Account Code","Account Name"] + month_cols)

    nm_src   = df[df["RowHadOwnCode"]] if "RowHadOwnCode" in df.columns else df
    name_map = (
        nm_src.dropna(subset=["Account Code"])[["Account Code","Account Name"]]
             .drop_duplicates("Account Code")
             .set_index("Account Code")["Account Name"]
    )

    totals_df = df[df.get("IsTotal", False)]
    codes_with_total = set(totals_df["Account Code"].dropna().astype(str).unique())

    cols = ["Account Code"] + [c for c in month_cols if c in df.columns]
    parts = []
    if not totals_df.empty:
        parts.append(totals_df[cols].groupby("Account Code", as_index=False).sum())

    no_total_df = df[~df["Account Code"].astype(str).isin(codes_with_total)]
    if not no_total_df.empty:
        parts.append(no_total_df[cols].groupby("Account Code", as_index=False).sum())

    agg = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame(columns=cols)
    agg["Account Name"] = agg["Account Code"].map(name_map).fillna("")
    return agg[["Account Code","Account Name"] + [c for c in month_cols if c in agg.columns]]

# -----------------------------------------------------------------------------
# MoM + trend signals
# -----------------------------------------------------------------------------

def compute_mom_with_trends(df: pd.DataFrame, month_cols: list[str], CONFIG: dict) -> pd.DataFrame:
    """Compute MoM deltas and a simple rolling average signal."""
    if len(month_cols) < 2:
        return pd.DataFrame(columns=[
            "Account Code","Account Name","Prior","Current","Delta","Pct Change","Period",
            "Trend_3M_Avg","Trend_Deviation"
        ])

    out: list[pd.DataFrame] = []
    for i in range(1, len(month_cols)):
        cur, prev = month_cols[i], month_cols[i-1]
        if cur not in df.columns or prev not in df.columns:
            continue

        tmp = df[["Account Code","Account Name", prev, cur]].copy()
        tmp = tmp.rename(columns={prev: "Prior", cur: "Current"})
        tmp["Delta"] = tmp["Current"] - tmp["Prior"]
        tmp["Pct Change"] = np.where(tmp["Prior"] == 0, np.nan, tmp["Delta"] / tmp["Prior"])
        tmp["Period"] = normalize_period_label(cur)

        # simple trend window using preceding up-to-5 periods (require min_trend_periods)
        if i >= CONFIG["min_trend_periods"]:
            start_idx = max(0, i - 5)
            trend_cols = month_cols[start_idx:i]
            if len(trend_cols) >= CONFIG["min_trend_periods"]:
                trend_data = df[trend_cols]
                tmp["Trend_3M_Avg"] = trend_data.mean(axis=1)
                tmp["Trend_Deviation"] = tmp["Current"] - tmp["Trend_3M_Avg"]
            else:
                tmp["Trend_3M_Avg"] = np.nan
                tmp["Trend_Deviation"] = np.nan
        else:
            tmp["Trend_3M_Avg"] = np.nan
            tmp["Trend_Deviation"] = np.nan

        out.append(tmp)

    return pd.concat(out, ignore_index=True) if out else pd.DataFrame(columns=[
        "Account Code","Account Name","Prior","Current","Delta","Pct Change","Period",
        "Trend_3M_Avg","Trend_Deviation"
    ])

# -----------------------------------------------------------------------------
# Rules & anomaly builders
# -----------------------------------------------------------------------------

def classify_pl_account(code: object, CONFIG: dict) -> str:
    """Return 'Recurring' if code starts with any configured prefix; else 'Revenue/OPEX'."""
    code_str = str(code)
    return "Recurring" if any(code_str.startswith(p) for p in CONFIG["recurring_code_prefixes"]) else "Revenue/OPEX"

def get_threshold_cause(statement: str, code: object, CONFIG: dict) -> str:
    """Human-friendly cause suggestion based on statement type and code classification."""
    if statement == "BS":
        return "Balance changed materially — check reclass/missing offset."
    return ("Recurring moved — check accruals/timing."
            if classify_pl_account(code, CONFIG) == "Recurring"
            else "Revenue/OPEX moved — check billing/cut-off.")

def match_codes(series: pd.Series, pattern_str: str | float | int | None) -> pd.Series:
    """Return boolean mask for 'code patterns' like '111*,112*|515'."""
    if pd.isna(pattern_str) or pattern_str == "":
        return pd.Series(False, index=series.index)
    patterns = [p.strip() for p in str(pattern_str).split("|") if p.strip()]
    mask = pd.Series(False, index=series.index)
    for pattern in patterns:
        if pattern.endswith("*"):
            prefix = pattern[:-1]
            mask |= series.astype(str).str.startswith(prefix)
        else:
            mask |= (series.astype(str) == pattern)
    return mask

def build_corr_anoms(
    sub: str,
    combined: pd.DataFrame,
    corr_rules: pd.DataFrame,
    periods: list[str],
    materiality: int,
) -> list[dict]:
    """Correlation-rule anomalies; flags when left/right deltas break expected direction."""
    items: list[dict] = []

    cols = {c.lower(): c for c in corr_rules.columns}
    def pick(opts: list[str]) -> str | None:
        for n in opts:
            if n in cols:
                return cols[n]
        return None

    left_col  = pick(["left_patterns","left_pattern","left_patter","left"])
    right_col = pick(["right_patterns","right_pattern","right_patter","right"])
    cause_col = pick(["cause_message","cause","message","notes"])
    type_col  = pick(["relation_type","type","direction"])

    for _, rule in corr_rules.iterrows():
        lp = str(rule[left_col]) if left_col else ""
        rp = str(rule[right_col]) if right_col else ""
        cause = str(rule[cause_col]) if cause_col else "Correlation mismatch"
        rel = str(rule.get(type_col, "directional")).strip().lower() if type_col else "directional"
        inverse = rel in ("inverse","opposite","neg","negative")
        if not lp or not rp:
            continue

        for per in periods:
            mom = combined[combined["Norm_Period"] == per]
            if mom.empty:
                continue

            l = mom[match_codes(mom["Account Code"], lp)]["Delta"].sum()
            r = mom[match_codes(mom["Account Code"], rp)]["Delta"].sum()

            ok = ((l > 0 and r < 0) or (l < 0 and r > 0)) if inverse else ((l > 0 and r > 0) or (l < 0 and r < 0))
            if abs(l) >= materiality and not ok:
                items.append({
                    "Subsidiary": sub,
                    "Account": f"{lp} ↔ {rp}",
                    "Period": per,
                    "Pct Change": "",
                    "Abs Change (VND)": int(l),
                    "Trigger(s)": "Correlation break",
                    "Suggested likely cause": cause,
                    "Status": "Needs Review",
                    "Notes": f"Left Δ={int(l):,}, Right Δ={int(r):,}, relation={'inverse' if inverse else 'directional'}",
                })
    return items

def build_gross_margin_anoms(
    sub: str,
    pl_data: pd.DataFrame,
    pl_cols: list[str],
    CONFIG: dict,
) -> list[dict]:
    """Gross margin % = (Revenue(511*) - COGS(632*)) / Revenue(511*).
    Flag when current GM% drops vs prior by >= gm_drop_threshold_pct (absolute).
    """
    items: list[dict] = []
    if pl_data is None or pl_data.empty or len(pl_cols) < 2:
        return items

    # Sum revenue & COGS by month across all 511* / 632* rows
    is_rev = pl_data["Account Code"].astype(str).str.startswith("511")
    is_cogs = pl_data["Account Code"].astype(str).str.startswith("632")

    rev = pl_data.loc[is_rev, pl_cols].sum(numeric_only=True)
    cogs = pl_data.loc[is_cogs, pl_cols].sum(numeric_only=True)

    # Walk month to month
    for i in range(1, len(pl_cols)):
        prev, cur = pl_cols[i-1], pl_cols[i]
        rev_prev, rev_cur = float(rev.get(prev, 0.0) or 0.0), float(rev.get(cur, 0.0) or 0.0)
        cogs_prev, cogs_cur = float(cogs.get(prev, 0.0) or 0.0), float(cogs.get(cur, 0.0) or 0.0)

        if rev_prev == 0 or rev_cur == 0:
            continue  # cannot compute a meaningful margin

        gm_prev = (rev_prev - cogs_prev) / rev_prev
        gm_cur  = (rev_cur  - cogs_cur)  / rev_cur
        drop = gm_cur - gm_prev  # negative when margin worsens

        if drop <= -abs(CONFIG.get("gm_drop_threshold_pct", 0.01)):
            items.append({
                "Subsidiary": sub,
                "Account": "Gross Margin (511-632)",
                "Period": normalize_period_label(cur),
                "Pct Change": round((gm_cur - gm_prev) * 100, 2),
                "Abs Change (VND)": "",
                "Trigger(s)": f"Gross margin drop ≥ {int(CONFIG.get('gm_drop_threshold_pct', 0.01)*100)}%",
                "Suggested likely cause": "COGS moved vs revenue; check pricing, mix, or timing.",
                "Status": "Needs Review",
                "Notes": f"GM {normalize_period_label(prev)}={gm_prev:.2%} → {normalize_period_label(cur)}={gm_cur:.2%}",
            })

    return items

def build_revenue_by_customer_anoms(
    sub: str,
    pl_data: pd.DataFrame,
    pl_cols: list[str],
    CONFIG: dict,
) -> list[dict]:
    """If pl_data includes a customer column, compute MoM by customer for 511* rows."""
    items: list[dict] = []
    if pl_data is None or pl_data.empty or len(pl_cols) < 2:
        return items

    cust_col = find_customer_column(pl_data, CONFIG)
    if not cust_col or cust_col not in pl_data.columns:
        return items  # nothing to do

    # Filter revenue rows
    rev_df = pl_data[pl_data["Account Code"].astype(str).str.startswith("511")].copy()
    if rev_df.empty:
        return items

    # Group by customer per month; then compute MoM % and abs deltas
    g = rev_df.groupby(cust_col)[pl_cols].sum(numeric_only=True)
    customers = g.index.tolist()

    for cust in customers:
        series = g.loc[cust]
        for i in range(1, len(pl_cols)):
            prev, cur = pl_cols[i-1], pl_cols[i]
            prev_v = float(series.get(prev, 0.0) or 0.0)
            cur_v  = float(series.get(cur, 0.0) or 0.0)
            delta  = cur_v - prev_v
            pct    = (delta / prev_v) if prev_v != 0 else np.nan

            # Use the same Revenue/OPEX rule: % > threshold OR abs ≥ materiality
            cond_pct = (pd.notna(pct) and abs(pct) > CONFIG["revenue_opex_pct_threshold"])
            cond_abs = (abs(delta) >= CONFIG["materiality_vnd"])
            if cond_pct or cond_abs:
                items.append({
                    "Subsidiary": sub,
                    "Account": f"Revenue 511* — Customer: {cust}",
                    "Period": normalize_period_label(cur),
                    "Pct Change": round(pct*100, 2) if pd.notna(pct) else "",
                    "Abs Change (VND)": int(delta),
                    "Trigger(s)": "Revenue by customer variance",
                    "Suggested likely cause": "Customer-level shift; check orders, churn, or timing.",
                    "Status": "Needs Review",
                    "Notes": f"{normalize_period_label(prev)}={int(prev_v):,} → {normalize_period_label(cur)}={int(cur_v):,}",
                })

    return items

# -----------------------------------------------------------------------------
# AI Analysis Integration
# -----------------------------------------------------------------------------

def build_anoms_ai_mode(
    sub: str,
    excel_bytes: bytes,
    filename: str,
    CONFIG: dict,
) -> pd.DataFrame:
    """AI-only anomaly detection using complete raw Excel file."""
    anomalies: list[dict] = []

    if not AI_AVAILABLE:
        print(f"\n❌ AI analysis requested but AI analyzer not available for '{sub}'")
        return pd.DataFrame([{
            "Subsidiary": sub,
            "Account": "AI_NOT_AVAILABLE",
            "Period": "N/A",
            "Pct Change": 0,
            "Abs Change (VND)": 0,
            "Trigger(s)": "AI Analysis Not Available",
            "Suggested likely cause": "AI analyzer module not found - install required dependencies",
            "Status": "Error",
            "Notes": "Check if llm_analyzer.py is available and dependencies are installed",
        }])

    print(f"\n🧠 Starting AI analysis for '{sub}'...")
    try:
        llm_analyzer = LLMFinancialAnalyzer(CONFIG.get("llm_model", "gpt-4o"))
        print(f"✅ AI analyzer initialized with model: {CONFIG.get('llm_model', 'gpt-4o')}")

        print(f"\n🔍 Running AI analysis on complete raw Excel file...")
        llm_anomalies = llm_analyzer.analyze_raw_excel_file(excel_bytes, filename, sub, CONFIG)
        print(f"✅ AI analysis completed, processing {len(llm_anomalies)} results")

        print(f"\n📋 Converting {len(llm_anomalies)} AI results to report format...")
        for idx, anom in enumerate(llm_anomalies, 1):
            print(f"   • Processing anomaly {idx}: Account {anom.get('account_code', 'Unknown')}")
            anomalies.append({
                "Subsidiary": anom["subsidiary"],
                "Account": anom["account_code"],
                "Period": "Current",
                "Pct Change": anom["change_percent"],
                "Abs Change (VND)": int(anom["change_amount"]),
                "Trigger(s)": anom["rule_name"],
                "Suggested likely cause": anom["details"],
                "Status": "AI Analysis",
                "Notes": anom["details"],
            })
        print(f"✅ Successfully converted all AI results to report format")

        print(f"\n✅ AI anomaly detection completed for '{sub}' - returning {len(anomalies)} records")
        return pd.DataFrame(anomalies)

    except Exception as e:
        print(f"\n❌ AI analysis failed for '{sub}': {e}")
        error_record = pd.DataFrame([{
            "Subsidiary": sub,
            "Account": "AI_ERROR",
            "Period": "N/A",
            "Pct Change": 0,
            "Abs Change (VND)": 0,
            "Trigger(s)": "AI Analysis Failed",
            "Suggested likely cause": f"AI model error: {str(e)[:100]}...",
            "Status": "Error",
            "Notes": "Check if OpenAI is running and model is available",
        }])
        print(f"⚠️  Returning error record to continue processing other files")
        return error_record

# -----------------------------------------------------------------------------
# Python Analysis Mode (Traditional Rule-Based)
# -----------------------------------------------------------------------------

def build_anoms_python_mode(
    sub: str,
    bs_data: pd.DataFrame, bs_cols: list[str],
    pl_data: pd.DataFrame, pl_cols: list[str],
    corr_rules: pd.DataFrame, season_rules: pd.DataFrame,
    CONFIG: dict,
) -> pd.DataFrame:
    """Apply all anomaly rules to one subsidiary's cleaned BS/PL frames."""
    anomalies: list[dict] = []
    materiality = CONFIG["materiality_vnd"]

    bs_mom = compute_mom_with_trends(bs_data, bs_cols, CONFIG)
    pl_mom = compute_mom_with_trends(pl_data, pl_cols, CONFIG)

    # Balance Sheet rule
    for _, row in bs_mom.iterrows():
        abs_delta = abs(row["Delta"])
        pct_change = row["Pct Change"]
        if (abs_delta >= materiality and pd.notna(pct_change) and abs(pct_change) > CONFIG["bs_pct_threshold"]):
            anomalies.append({
                "Subsidiary": sub,
                "Account": f'{row["Account Code"]}-{row["Account Name"]}',
                "Period": row["Period"],
                "Pct Change": round(pct_change * 100, 2),
                "Abs Change (VND)": int(row["Delta"]),
                "Trigger(s)": "BS >5% & ≥1B",
                "Suggested likely cause": get_threshold_cause("BS", row["Account Code"], CONFIG),
                "Status": "Needs Review",
                "Notes": "",
            })

    # P/L rules split: Recurring vs Revenue/OPEX (+ special %-only for dep accounts)
    dep_prefixes = [str(p) for p in CONFIG.get("dep_pct_only_prefixes", [])]

    def is_dep_pct_only(code: object) -> bool:
        cs = str(code)
        return any(cs.startswith(p) for p in dep_prefixes)

    for _, row in pl_mom.iterrows():
        abs_delta = abs(row["Delta"])
        pct_change = row["Pct Change"]
        code = row["Account Code"]
        account_class = classify_pl_account(code, CONFIG)
        trigger = ""

        if is_dep_pct_only(code):
            if (pd.notna(pct_change) and abs(pct_change) > CONFIG["recurring_pct_threshold"]):
                trigger = "Depreciation % change > threshold"
        elif account_class == "Recurring":
            if (abs_delta >= materiality and pd.notna(pct_change) and abs(pct_change) > CONFIG["recurring_pct_threshold"]):
                trigger = "Recurring >5% & ≥1B"
        else:
            if ((pd.notna(pct_change) and abs(pct_change) > CONFIG["revenue_opex_pct_threshold"]) or abs_delta >= materiality):
                trigger = "Revenue/OPEX >10% or ≥1B"

        if trigger:
            anomalies.append({
                "Subsidiary": sub,
                "Account": f'{row["Account Code"]}-{row["Account Name"]}',
                "Period": row["Period"],
                "Pct Change": round(pct_change * 100, 2) if pd.notna(pct_change) else "",
                "Abs Change (VND)": int(row["Delta"]),
                "Trigger(s)": trigger,
                "Suggested likely cause": get_threshold_cause("PL", row["Account Code"], CONFIG),
                "Status": "Needs Review",
                "Notes": "",
            })

    # Gross Margin anomalies
    anomalies.extend(build_gross_margin_anoms(sub, pl_data, pl_cols, CONFIG))

    # Revenue-by-customer anomalies (if a customer column exists)
    anomalies.extend(build_revenue_by_customer_anoms(sub, pl_data, pl_cols, CONFIG))

    # Correlation rules (optional)
    combined = pd.concat([
        bs_mom[["Account Code","Period","Delta"]],
        pl_mom[["Account Code","Period","Delta"]],
    ], ignore_index=True)
    combined["Norm_Period"] = combined["Period"].astype(str).map(normalize_period_label)
    periods = sorted(set(combined["Norm_Period"]), key=month_key)
    anomalies.extend(build_corr_anoms(sub, combined, corr_rules, periods, materiality))

    # Accounting-focused anomalies (wrapper)
    acct_anoms_df = run_accounting_rules_on_frames(pl_data, bs_data, subsidiary=sub)

    # Build final DataFrame
    main_df = pd.DataFrame(anomalies)
    if acct_anoms_df is not None and not acct_anoms_df.empty:
        if not main_df.empty:
            main_df = pd.concat([main_df, acct_anoms_df], ignore_index=True)
        else:
            main_df = acct_anoms_df

    return main_df

# -----------------------------------------------------------------------------
# Accounting-specific analysis functions
# -----------------------------------------------------------------------------

def check_gross_margin(pl_pivot: pd.DataFrame, entity_col: str = "Entity") -> list[dict]:
    """
    Needs rows for 511* (Revenue) and 632* (COGS) per entity. pl_pivot index should let us filter by account prefix.
    Returns anomaly records where gross margin % moves by >= threshold m/m.
    """
    out = []
    months = _months(pl_pivot)
    if not months:
        return out

    # group by entity
    for entity, dfE in pl_pivot.groupby(entity_col):
        # total 511 and 632 for the entity
        rev = dfE[dfE["Account"].apply(_is_511)][months].sum(numeric_only=True)
        cogs = dfE[dfE["Account"].apply(_is_632)][months].sum(numeric_only=True)

        # margin %
        with np.errstate(divide='ignore', invalid='ignore'):
            margin = (rev - cogs) / rev.replace({0: np.nan})

        # m/m change in percentage points
        for i in range(1, len(months)):
            m, pm = months[i], months[i-1]
            if pd.notna(margin[m]) and pd.notna(margin[pm]):
                delta = float(margin[m] - margin[pm])  # absolute change in fraction
                if abs(delta) >= ACCT_THRESH["gross_margin_pct_delta"]:
                    out.append(_anom_record(
                        "GrossMargin Δ≥1pp",
                        entity, "Gross Margin", m,
                        value=float(margin[m]),
                        detail={
                            "PrevMonth": pm,
                            "PrevValue": float(margin[pm]),
                            "DeltaPctPoints": round(delta*100, 2)
                        }
                    ))
    return out

def check_depreciation_variance(pl_pivot: pd.DataFrame, bs_pivot: pd.DataFrame, entity_col: str = "Entity") -> list[dict]:
    """
    Track % changes for 217*, 632* depreciation, and 214 (SCC).
    - P&L: 632* lines that are depreciation (we treat all 632* for simplicity; refine if you have sub-accounts)
    - BS: 217* and 214* balances delta month-over-month
    """
    out = []
    months = _months(pl_pivot)
    if not months:
        return out

    for entity, dfE in pl_pivot.groupby(entity_col):
        # P&L depreciation proxy: total 632*
        pl_dep = dfE[dfE["Account"].apply(_is_632)][months].sum(numeric_only=True)
        pl_pct = _pct_change(pl_dep)

        for i, m in enumerate(months):
            if i == 0:
                continue
            if pd.notna(pl_pct[m]) and abs(pl_pct[m]) >= ACCT_THRESH["depr_pct_delta"]:
                out.append(_anom_record(
                    "P&L Depreciation Δ%",
                    entity, "632* (Depreciation proxy)", m,
                    value=float(pl_dep[m]),
                    detail={"PctChange": float(pl_pct[m])}
                ))

        # BS 217*
        if entity in bs_pivot[entity_col].values:
            dfB = bs_pivot[bs_pivot[entity_col] == entity]
        else:
            dfB = bs_pivot  # fallback if bs not entity-split

        bal_217 = dfB[dfB["Account"].apply(_is_217)][months].sum(numeric_only=True)
        bal_217_pct = _pct_change(bal_217)
        for i, m in enumerate(months):
            if i == 0:
                continue
            if pd.notna(bal_217_pct[m]) and abs(bal_217_pct[m]) >= ACCT_THRESH["depr_pct_delta"]:
                out.append(_anom_record(
                    "BS 217* Δ%",
                    entity, "217* (Acc. Depreciation)", m,
                    value=float(bal_217[m]),
                    detail={"PctChange": float(bal_217_pct[m])}
                ))

        # BS 214* (esp. SCC)
        bal_214 = dfB[dfB["Account"].apply(_is_214)][months].sum(numeric_only=True)
        if not bal_214.empty:
            bal_214_pct = _pct_change(bal_214)
            for i, m in enumerate(months):
                if i == 0:
                    continue
                if pd.notna(bal_214_pct[m]) and abs(bal_214_pct[m]) >= ACCT_THRESH["depr_pct_delta"]:
                    out.append(_anom_record(
                        "BS 214* Δ%",
                        entity, "214* (Acc. Depreciation SCC)", m,
                        value=float(bal_214[m]),
                        detail={"PctChange": float(bal_214_pct[m])}
                    ))
    return out

def check_cogs_vs_revenue_ratio(pl_pivot: pd.DataFrame, entity_col: str = "Entity") -> list[dict]:
    """
    COGS/Revenue ratio drift > threshold vs historical mean.
    """
    out = []
    months = _months(pl_pivot)
    if not months:
        return out

    for entity, dfE in pl_pivot.groupby(entity_col):
        rev = dfE[dfE["Account"].apply(_is_511)][months].sum(numeric_only=True)
        cogs = dfE[dfE["Account"].apply(_is_632)][months].sum(numeric_only=True)
        ratio = (cogs / rev.replace({0: np.nan})).astype(float)

        hist_mean = float(ratio.mean(skipna=True)) if ratio.notna().any() else np.nan
        if pd.isna(hist_mean):
            continue

        for m in months:
            if pd.notna(ratio[m]) and abs(ratio[m] - hist_mean) >= ACCT_THRESH["cogs_ratio_delta"]:
                out.append(_anom_record(
                    "COGS/Revenue ratio drift",
                    entity, "632* vs 511*", m,
                    value=float(ratio[m]),
                    detail={"HistMean": hist_mean, "Delta": float(ratio[m] - hist_mean)}
                ))
    return out

def check_sga_as_pct_of_revenue(pl_pivot: pd.DataFrame, entity_col: str = "Entity") -> list[dict]:
    """
    SG&A as % of revenue exceeding historical mean by > threshold.
    """
    out = []
    months = _months(pl_pivot)
    if not months:
        return out

    for entity, dfE in pl_pivot.groupby(entity_col):
        rev = dfE[dfE["Account"].apply(_is_511)][months].sum(numeric_only=True)
        sga = dfE[dfE["Account"].apply(lambda a: _is_641(a) or _is_642(a))][months].sum(numeric_only=True)

        mean_pct, std_pct = _series_hist_pct_of_rev(sga, rev)
        if pd.isna(mean_pct):
            continue

        # Flag if current % > mean + 10% points of mean (relative), or simply > mean*(1+delta)
        for m in months:
            if rev.get(m, np.nan) == 0 or pd.isna(rev.get(m, np.nan)):
                continue
            pct = float(sga.get(m, np.nan) / rev.get(m, np.nan))
            if pd.notna(pct) and pct > (mean_pct * (1 + ACCT_THRESH["sga_pct_of_rev_delta"])):
                out.append(_anom_record(
                    "SG&A % of Revenue spike",
                    entity, "641*/642*", m,
                    value=float(sga.get(m, np.nan)),
                    detail={"PctOfRevenue": pct, "HistMean": mean_pct}
                ))
    return out

def check_financial_items_swings(pl_pivot: pd.DataFrame, entity_col: str = "Entity") -> list[dict]:
    """
    Financial expenses (635*) and income (515*) percentage swings > threshold.
    """
    out = []
    months = _months(pl_pivot)
    if not months:
        return out

    for entity, dfE in pl_pivot.groupby(entity_col):
        for prefix, label, pred in [
            ("635*", "Financial expenses (635*)", _is_635),
            ("515*", "Financial income (515*)", _is_515),
        ]:
            series = dfE[dfE["Account"].apply(pred)][months].sum(numeric_only=True)
            pct = _pct_change(series)
            for i, m in enumerate(months):
                if i == 0:
                    continue
                if pd.notna(pct[m]) and abs(pct[m]) >= ACCT_THRESH["fin_swing_pct"]:
                    out.append(_anom_record(
                        f"{label} swing",
                        entity, prefix, m,
                        value=float(series[m]),
                        detail={"PctChange": float(pct[m])}
                    ))
    return out

def check_bs_pl_dep_consistency(pl_pivot: pd.DataFrame, bs_pivot: pd.DataFrame, entity_col: str = "Entity") -> list[dict]:
    """
    Compare BS accumulated depreciation (217* + 214*) Δ vs P&L 632* depreciation expense.
    Flag when mismatch > threshold %.
    """
    out = []
    months = _months(pl_pivot)
    if not months:
        return out

    for entity, dfE in pl_pivot.groupby(entity_col):
        # P&L dep expense proxy
        pl_dep = dfE[dfE["Account"].apply(_is_632)][months].sum(numeric_only=True)

        if entity in bs_pivot[entity_col].values:
            dfB = bs_pivot[bs_pivot[entity_col] == entity]
        else:
            dfB = bs_pivot

        bs_acc_dep = dfB[dfB["Account"].apply(lambda a: _is_217(a) or _is_214(a))][months].sum(numeric_only=True)

        # month-over-month deltas
        pl_delta = pl_dep.diff()
        bs_delta = bs_acc_dep.diff()

        for m in months:
            if m == months[0]:
                continue
            x = float(pl_delta.get(m, np.nan)) if pd.notna(pl_delta.get(m, np.nan)) else None
            y = float(bs_delta.get(m, np.nan)) if pd.notna(bs_delta.get(m, np.nan)) else None
            if x is None or y is None or y == 0:
                continue
            diff_pct = abs(x - y) / (abs(y) if y != 0 else np.nan)
            if pd.notna(diff_pct) and diff_pct > ACCT_THRESH["bs_pl_dep_diff_pct"]:
                out.append(_anom_record(
                    "BS↔PL Depreciation mismatch",
                    entity, "217*+214* vs 632*", m,
                    value={"PL_Dep_Delta": x, "BS_AccDep_Delta": y},
                    detail={"DiffPct": float(diff_pct)}
                ))
    return out

# -----------------------------------------------------------------------------
# File-level processing (IN-MEMORY)
# -----------------------------------------------------------------------------

def process_financial_tab_from_bytes(
    xl_bytes: bytes,
    sheet_name: str,
    mode: str,
    subsidiary: str,
) -> tuple[pd.DataFrame, list[str]]:
    """Load and clean one sheet ('BS Breakdown' or 'PL Breakdown') from in-memory bytes."""
    header_row = detect_header_row(xl_bytes, sheet_name)
    df = pd.read_excel(io.BytesIO(xl_bytes), sheet_name=sheet_name, header=header_row, dtype=str)
    df = normalize_financial_col(df)
    df, month_cols = promote_row8(df, mode, subsidiary)
    df = fill_down_assign(df)
    df = coerce_numeric(df, month_cols)
    keep_cols = ["Account Code","Account Name","RowHadOwnCode","IsTotal"] + [c for c in month_cols if c in df.columns]
    df = df[keep_cols]
    totals = aggregate_totals(df, month_cols)
    return totals, month_cols

def extract_subsidiary_name_from_bytes(xl_bytes: bytes, fallback_filename: str) -> str:
    """Try to find a name on A2 of BS/PL sheets like 'Subsidiary: XYZ'. Fallback to filename stem."""
    try:
        wb = load_workbook(io.BytesIO(xl_bytes), read_only=True, data_only=True)
        for sheet_name in ["BS Breakdown", "PL Breakdown"]:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                cell_value = sheet["A2"].value
                if isinstance(cell_value, str) and ":" in cell_value:
                    wb.close()
                    return cell_value.split(":")[-1].strip()
        wb.close()
    except Exception:
        pass
    # fallback: filename before first underscore or dot
    stem = fallback_filename.rsplit("/", 1)[-1]
    stem = stem.split("\\")[-1]
    stem = stem.split(".")[0]
    return stem.split("_")[0] if "_" in stem else stem

# -----------------------------------------------------------------------------
# Excel formatting (IN-MEMORY, works on a worksheet not a saved file)
# -----------------------------------------------------------------------------

def apply_excel_formatting_ws(ws, anomaly_df: pd.DataFrame, CONFIG: dict) -> None:
    """Apply simple conditional fills directly on the 'Anomalies Summary' worksheet."""
    try:
        critical_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        warning_fill  = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        header_fill   = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        ai_fill       = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")

        # Header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)

        # Find indexes for columns we care about
        headers = [c.value for c in ws[1]]
        try:
            abs_idx = headers.index("Abs Change (VND)") + 1
            trig_idx = headers.index("Trigger(s)") + 1
            status_idx = headers.index("Status") + 1
        except ValueError:
            # If columns not found, skip formatting
            return

        # Rows
        for row_idx in range(2, ws.max_row + 1):
            try:
                abs_change = ws.cell(row=row_idx, column=abs_idx).value or 0
                trigger = str(ws.cell(row=row_idx, column=trig_idx).value or "")
                status = str(ws.cell(row=row_idx, column=status_idx).value or "")

                fill = None
                if "AI Analysis" in status:
                    # Light green for AI analysis
                    fill = ai_fill
                elif abs_change >= CONFIG.get("materiality_vnd", 1000000000) * 5:
                    fill = critical_fill
                elif "Correlation break" in trigger or abs_change >= CONFIG.get("materiality_vnd", 1000000000) * 2:
                    fill = warning_fill

                if fill:
                    for col_idx in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill
            except Exception:
                continue
    except Exception:
        # Formatting should never break
        pass

# -----------------------------------------------------------------------------
# Accounting-specific wrapper
# -----------------------------------------------------------------------------

def _month_cols(df: pd.DataFrame) -> list[str]:
    patt = re.compile(r"^[A-Za-z]{3}\s+\d{4}$")
    return [c for c in df.columns if isinstance(c, str) and patt.match(c)]

def _ensure_account_and_entity(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "Entity" not in out.columns:
        if "Subsidiary" in out.columns:
            out.rename(columns={"Subsidiary": "Entity"}, inplace=True)
        else:
            out["Entity"] = "ALL"
    if "Account" not in out.columns:
        # Fix: ensure "Financial row" matches the standardized casing
        for cand in ["Account (Line): Mã số - Code", "Account (Line)", "Account Code", "Financial row"]:
            if cand in out.columns:
                out["Account"] = out[cand].astype(str)
                break
        if "Account" not in out.columns:
            out["Account"] = out.iloc[:,0].astype(str)
    return out[["Entity","Account"] + _month_cols(out)].copy()

def _map_accounting_records_to_summary(acct_df: pd.DataFrame, subsidiary: str) -> pd.DataFrame:
    if acct_df.empty:
        return acct_df

    def _get_pct(row):
        for k in ["PctChange","DeltaPct","DeltaPctPoints","DiffPct"]:
            if k in row and pd.notna(row[k]):
                return float(row[k])
        return np.nan

    def _get_abs(row):
        v = row.get("Value", np.nan)
        if isinstance(v, dict):
            return np.nan
        try:
            return float(v)
        except Exception:
            return np.nan

    def _guess_cause(rule):
        rule = str(rule).lower()
        if "grossmargin" in rule:
            return "Pricing/COGS timing or one-off cost"
        if "depreciation" in rule:
            return "New asset, disposal, or misclassification"
        if "cogs/revenue" in rule:
            return "Inefficiency or pricing mismatch"
        if "sg&a" in rule:
            return "Overhead spike vs revenue"
        if "financial" in rule:
            return "Debt/investment swing or FX"
        if "dep mismatch" in rule:
            return "BS–PL inconsistency in depreciation"
        return ""

    rows = []
    for _, r in acct_df.iterrows():
        rows.append({
            "Subsidiary": subsidiary,
            "Account": r.get("Account",""),
            "Period": r.get("Month",""),
            "Pct Change": _get_pct(r),
            "Abs Change (VND)": _get_abs(r),
            "Trigger(s)": r.get("Rule",""),
            "Suggested likely cause": _guess_cause(r.get("Rule","")),
            "Status": "",
            "Notes": ""
        })
    return pd.DataFrame(rows)

def run_accounting_rules_on_frames(pl_df: pd.DataFrame, bs_df: pd.DataFrame, *, subsidiary: str) -> pd.DataFrame:
    if pl_df is None or pl_df.empty:
        return pd.DataFrame()

    pl_norm = _ensure_account_and_entity(pl_df)
    months = _month_cols(pl_norm)
    if not months:
        return pd.DataFrame()

    if bs_df is None or bs_df.empty:
        bs_norm = pd.DataFrame(columns=["Entity","Account"] + months)
    else:
        bs_norm = _ensure_account_and_entity(bs_df)

    acct_records = []
    acct_records += check_gross_margin(pl_norm)
    acct_records += check_depreciation_variance(pl_norm, bs_norm)
    acct_records += check_cogs_vs_revenue_ratio(pl_norm)
    acct_records += check_sga_as_pct_of_revenue(pl_norm)
    acct_records += check_financial_items_swings(pl_norm)
    acct_records += check_bs_pl_dep_consistency(pl_norm, bs_norm)

    acct_df = pd.DataFrame(acct_records)
    return _map_accounting_records_to_summary(acct_df, subsidiary)

# -----------------------------------------------------------------------------
# MAIN ENTRY POINT (UNIFIED)
# -----------------------------------------------------------------------------

def process_all(
    files: list[tuple[str, bytes]],
    corr_rules: Optional[pd.DataFrame] = None,
    season_rules: Optional[pd.DataFrame] = None,
    CONFIG: dict = DEFAULT_CONFIG,
    progress_callback=None
) -> bytes | tuple[bytes, list[tuple[str, bytes]]]:
    """
    Unified processing function that handles both Python and AI analysis modes.

    Args:
        files: List of (filename, bytes) tuples
        corr_rules: Correlation rules DataFrame (Python mode only)
        season_rules: Seasonality rules DataFrame (Python mode only)
        CONFIG: Configuration dictionary
        progress_callback: Progress callback function (AI mode only)

    Returns:
        bytes: Excel file bytes (Python mode)
        tuple[bytes, list[tuple[str, bytes]]]: Excel bytes and debug files (AI mode)
    """
    use_ai = CONFIG.get("use_llm_analysis", False)

    if use_ai:
        return process_all_ai_mode(files, CONFIG, progress_callback)
    else:
        return process_all_python_mode(files, corr_rules, season_rules, CONFIG)

def _add_revenue_analysis_to_sheet(ws, revenue_analysis: dict):
    """Add revenue analysis data to an Excel worksheet in a structured format."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(bold=True, size=12, color="2F5597")
    currency_font = Font(name="Arial", size=10)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def format_vnd(amount):
        if isinstance(amount, (int, float)) and not pd.isna(amount):
            return f"{amount:,.0f} VND"
        return "N/A"

    row = 1

    # Title
    ws[f"A{row}"] = "COMPREHENSIVE REVENUE ANALYSIS"
    ws[f"A{row}"].font = Font(bold=True, size=16, color="2F5597")
    row += 2

    # Executive Summary
    if revenue_analysis.get('summary'):
        summary = revenue_analysis['summary']
        ws[f"A{row}"] = "EXECUTIVE SUMMARY"
        ws[f"A{row}"].font = section_font
        row += 1

        summary_data = [
            ["Subsidiary", revenue_analysis.get('subsidiary', 'N/A')],
            ["Months Analyzed", len(revenue_analysis.get('months_analyzed', []))],
            ["Revenue Accounts", summary.get('total_accounts', 0)],
            ["Latest Total Revenue", format_vnd(summary.get('total_revenue_latest', 0))],
            ["Latest Gross Margin %", f"{summary.get('gross_margin_latest', 0):.1f}%" if summary.get('gross_margin_latest') else 'N/A']
        ]

        for label, value in summary_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
        row += 1

    # Risk Assessment
    if revenue_analysis.get('risk_assessment'):
        ws[f"A{row}"] = "RISK ASSESSMENT"
        ws[f"A{row}"].font = section_font
        row += 1

        # Headers
        headers = ["Period", "Risk Level", "Type", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        for risk in revenue_analysis['risk_assessment']:
            ws[f"A{row}"] = risk.get('period', '')
            ws[f"B{row}"] = risk.get('risk_level', '')
            ws[f"C{row}"] = risk.get('type', '')
            ws[f"D{row}"] = risk.get('description', '')

            # Color code risk levels
            risk_level = risk.get('risk_level', '').lower()
            if risk_level == 'high':
                fill_color = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            elif risk_level == 'medium':
                fill_color = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            else:
                fill_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")

            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill_color
                cell.border = thin_border
            row += 1
        row += 1

    # Total Revenue Trend
    if revenue_analysis.get('total_revenue_analysis', {}).get('changes'):
        ws[f"A{row}"] = "TOTAL REVENUE TREND (511*)"
        ws[f"A{row}"].font = section_font
        row += 1

        # Headers
        headers = ["Period", "Previous Value", "Current Value", "Change (VND)", "Change (%)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        for change in revenue_analysis['total_revenue_analysis']['changes']:
            period = f"{change.get('from', '')} → {change.get('to', '')}"
            ws[f"A{row}"] = period
            ws[f"B{row}"] = format_vnd(change.get('prev_value', 0))
            ws[f"C{row}"] = format_vnd(change.get('curr_value', 0))
            ws[f"D{row}"] = format_vnd(change.get('change', 0))
            ws[f"E{row}"] = f"{change.get('pct_change', 0):+.1f}%"

            # Color code positive/negative changes
            change_val = change.get('change', 0)
            if change_val > 0:
                fill_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            elif change_val < 0:
                fill_color = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            else:
                fill_color = None

            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                if fill_color:
                    cell.fill = fill_color
                cell.border = thin_border
            row += 1
        row += 1

    # Revenue by Account & Customer Impact Analysis - DETAILED BREAKDOWN
    if revenue_analysis.get('revenue_by_account'):
        ws[f"A{row}"] = "REVENUE BY ACCOUNT & CUSTOMER IMPACT ANALYSIS"
        ws[f"A{row}"].font = section_font
        row += 1

        for account_name, account_data in revenue_analysis['revenue_by_account'].items():
            if account_data.get('biggest_change') and abs(account_data['biggest_change'].get('change', 0)) > 1000000:
                # Account header
                ws[f"A{row}"] = f"Account: {account_name}"
                ws[f"A{row}"].font = Font(bold=True, size=11, color="1F4788")
                row += 1

                # Account biggest change
                biggest_change = account_data['biggest_change']
                ws[f"A{row}"] = "Biggest Change Period"
                ws[f"B{row}"] = f"{biggest_change.get('from', '')} → {biggest_change.get('to', '')}"
                ws[f"A{row}"].font = Font(bold=True)
                row += 1

                ws[f"A{row}"] = "Change Amount"
                ws[f"B{row}"] = format_vnd(biggest_change.get('change', 0))
                ws[f"A{row}"].font = Font(bold=True)
                row += 1

                ws[f"A{row}"] = "Change Percentage"
                ws[f"B{row}"] = f"{biggest_change.get('pct_change', 0):+.1f}%"
                ws[f"A{row}"].font = Font(bold=True)
                row += 1

                # Customer impacts table
                if account_data.get('customer_impacts'):
                    row += 1
                    ws[f"A{row}"] = "TOP CUSTOMER IMPACTS:"
                    ws[f"A{row}"].font = Font(bold=True, color="8B0000")
                    row += 1

                    # Customer impact headers
                    impact_headers = ["Customer/Entity", "Previous Value", "Current Value", "Change (VND)", "Change (%)"]
                    for col, header in enumerate(impact_headers, 1):
                        cell = ws.cell(row=row, column=col, value=header)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                        cell.border = thin_border
                    row += 1

                    # Customer impact data
                    for impact in account_data['customer_impacts']:
                        ws[f"A{row}"] = impact.get('entity', '')
                        ws[f"B{row}"] = format_vnd(impact.get('prev_val', 0))
                        ws[f"C{row}"] = format_vnd(impact.get('curr_val', 0))
                        ws[f"D{row}"] = format_vnd(impact.get('change', 0))
                        ws[f"E{row}"] = f"{impact.get('pct_change', 0):+.1f}%"

                        # Color code positive/negative changes
                        change_val = impact.get('change', 0)
                        if change_val > 0:
                            fill_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                        elif change_val < 0:
                            fill_color = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                        else:
                            fill_color = None

                        for col in range(1, 6):
                            cell = ws.cell(row=row, column=col)
                            if fill_color:
                                cell.fill = fill_color
                            cell.border = thin_border
                        row += 1

                row += 2  # Space between accounts

    # Gross Margin Analysis
    if revenue_analysis.get('gross_margin_analysis', {}).get('trend'):
        ws[f"A{row}"] = "GROSS MARGIN ANALYSIS"
        ws[f"A{row}"].font = section_font
        row += 1

        # Headers
        headers = ["Month", "Revenue", "Cost", "Gross Margin %", "Change from Previous"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        prev_margin = None
        for margin_data in revenue_analysis['gross_margin_analysis']['trend']:
            ws[f"A{row}"] = margin_data.get('month', '')
            ws[f"B{row}"] = format_vnd(margin_data.get('revenue', 0))
            ws[f"C{row}"] = format_vnd(margin_data.get('cost', 0))
            ws[f"D{row}"] = f"{margin_data.get('gross_margin_pct', 0):.1f}%"

            # Calculate change from previous month
            current_margin = margin_data.get('gross_margin_pct', 0)
            if prev_margin is not None:
                change = current_margin - prev_margin
                ws[f"E{row}"] = f"{change:+.1f}pp"
                # Color code margin changes
                if change > 0:
                    ws.cell(row=row, column=5).fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                elif change < 0:
                    ws.cell(row=row, column=5).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            else:
                ws[f"E{row}"] = "N/A"

            prev_margin = current_margin

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = thin_border
            row += 1
        row += 1

    # Utility Analysis (if available)
    if revenue_analysis.get('utility_analysis'):
        ws[f"A{row}"] = "UTILITY REVENUE VS COST ANALYSIS"
        ws[f"A{row}"].font = section_font
        row += 1

        if revenue_analysis['utility_analysis'].get('available') and revenue_analysis['utility_analysis'].get('margins'):
            # Headers
            headers = ["Month", "Utility Revenue", "Utility Cost", "Margin (VND)", "Margin %"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            row += 1

            for margin in revenue_analysis['utility_analysis']['margins']:
                ws[f"A{row}"] = margin.get('month', '')
                ws[f"B{row}"] = format_vnd(margin.get('revenue', 0))
                ws[f"C{row}"] = format_vnd(margin.get('cost', 0))
                ws[f"D{row}"] = format_vnd(margin.get('revenue', 0) - margin.get('cost', 0))
                ws[f"E{row}"] = f"{margin.get('margin_pct', 0):.1f}%"

                # Color code utility margins
                margin_pct = margin.get('margin_pct', 0)
                if margin_pct >= 0:
                    fill_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                else:
                    fill_color = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill_color
                    cell.border = thin_border
                row += 1
        else:
            ws[f"A{row}"] = "Utility accounts not found in the data."
            ws[f"A{row}"].font = Font(italic=True, color="666666")
            row += 1
        row += 1

    # Auto-fit columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def process_all_python_mode(
    files: list[tuple[str, bytes]],
    corr_rules: Optional[pd.DataFrame] = None,
    season_rules: Optional[pd.DataFrame] = None,
    CONFIG: dict = DEFAULT_CONFIG
) -> bytes:
    """Python rule-based analysis mode."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Anomalies Summary"
    all_anoms: list[pd.DataFrame] = []

    # default empty rules if None
    corr_rules = corr_rules if corr_rules is not None else pd.DataFrame()
    season_rules = season_rules if season_rules is not None else pd.DataFrame()

    for fname, xl_bytes in files:
        sub = extract_subsidiary_name_from_bytes(xl_bytes, fname)

        # Be forgiving if a sheet is missing
        bs_df, bs_cols = pd.DataFrame(), []
        pl_df, pl_cols = pd.DataFrame(), []
        try:
            bs_df, bs_cols = process_financial_tab_from_bytes(xl_bytes, "BS Breakdown", "BS", sub)
        except Exception:
            pass
        try:
            pl_df, pl_cols = process_financial_tab_from_bytes(xl_bytes, "PL Breakdown", "PL", sub)
        except Exception:
            pass

        anoms = build_anoms_python_mode(sub, bs_df, bs_cols, pl_df, pl_cols, corr_rules, season_rules, CONFIG)
        if anoms is not None and not anoms.empty:
            all_anoms.append(anoms)

    # Safe concat (even if no anomalies/files)
    if all_anoms:
        anom_df = pd.concat(all_anoms, ignore_index=True)
    else:
        anom_df = pd.DataFrame(columns=[
            "Subsidiary","Account","Period","Pct Change","Abs Change (VND)",
            "Trigger(s)","Suggested likely cause","Status","Notes"
        ])

    for r in dataframe_to_rows(anom_df, index=False, header=True):
        ws.append(r)
    apply_excel_formatting_ws(ws, anom_df, CONFIG)

    # === ADD REVENUE ANALYSIS SHEET ===
    print(f"\n📊 Adding Revenue Analysis sheet...")
    try:
        # Run revenue analysis for the first file
        if files:
            first_file_name, first_file_bytes = files[0]
            revenue_analysis = analyze_comprehensive_revenue_impact_from_bytes(first_file_bytes, first_file_name)

            # Create revenue analysis sheet
            rev_ws = wb.create_sheet(title="Revenue Analysis")
            _add_revenue_analysis_to_sheet(rev_ws, revenue_analysis)
            print(f"✅ Revenue Analysis sheet added successfully")
    except Exception as e:
        print(f"⚠️  Revenue Analysis sheet creation failed: {e}")
        # Continue without revenue analysis if it fails

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def process_all_ai_mode(
    files: list[tuple[str, bytes]],
    CONFIG: dict = DEFAULT_CONFIG,
    progress_callback=None
) -> tuple[bytes, list[tuple[str, bytes]]]:
    """AI-powered analysis mode."""
    print(f"\n🚀 ===== STARTING AI VARIANCE ANALYSIS PROCESSING =====\n")
    print(f"📥 Processing {len(files)} Excel file(s) for AI analysis")
    print(f"🤖 LLM Model: {CONFIG.get('llm_model', 'gpt-4o')}")
    print(f"🔧 AI-Only Mode: {CONFIG.get('use_llm_analysis', True)}")

    # === EXCEL WORKBOOK INITIALIZATION ===
    print(f"\n📊 Initializing Excel workbook for results...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Anomalies Summary"
    all_anoms: list[pd.DataFrame] = []
    debug_files: list[tuple[str, bytes]] = []  # Store debug files for download
    print(f"✅ Excel workbook initialized successfully")

    # === MULTI-FILE PROCESSING LOOP ===
    print(f"\n🔄 Starting processing loop for {len(files)} file(s)...\n")

    for file_idx, (fname, xl_bytes) in enumerate(files, 1):
        # Calculate progress range for this file (30% to 80% of total)
        file_start = 30 + ((file_idx - 1) * 50 // len(files))
        file_end = 30 + (file_idx * 50 // len(files))

        if progress_callback:
            progress_callback(file_start, f"Processing file {file_idx}/{len(files)}: {fname}")

        print(f"\n📁 ===== PROCESSING FILE {file_idx}/{len(files)} =====\n")
        print(f"📄 File: {fname}")
        print(f"📏 File Size: {len(xl_bytes):,} bytes ({len(xl_bytes)/1024:.1f} KB)")

        if progress_callback:
            progress_callback(file_start + 2, f"Extracting subsidiary name from {fname}")

        print(f"\n🏢 Extracting subsidiary name...")
        sub = extract_subsidiary_name_from_bytes(xl_bytes, fname)
        print(f"✅ Subsidiary: '{sub}'")

        if progress_callback:
            progress_callback(file_start + 5, f"Starting AI analysis for {sub}")

        # === AI ANALYSIS ===
        print(f"\n🤖 Starting AI analysis for '{sub}'...")
        anoms = build_anoms_ai_mode(sub, xl_bytes, fname, CONFIG)

        if progress_callback:
            progress_callback(file_end - 5, f"AI analysis complete for {sub}")

        if anoms is not None and not anoms.empty:
            print(f"✅ AI analysis completed successfully")
            print(f"   • Anomalies detected: {len(anoms)}")
            if len(anoms) > 0:
                ai_status_count = anoms['Status'].value_counts().to_dict()
                for status, count in ai_status_count.items():
                    print(f"   • {status}: {count}")
            all_anoms.append(anoms)
        else:
            print(f"⚠️  No anomalies detected or AI analysis returned empty result")

        print(f"\n✅ File '{fname}' processing completed\n")

    # === CONSOLIDATION & EXCEL GENERATION ===
    print(f"\n📊 ===== CONSOLIDATING RESULTS =====\n")
    print(f"📈 Processed {len(files)} file(s) successfully")

    if all_anoms:
        print(f"🔗 Consolidating {len(all_anoms)} result set(s)...")
        anom_df = pd.concat(all_anoms, ignore_index=True)
        print(f"✅ Consolidation completed")
        print(f"   • Total anomalies: {len(anom_df)}")

        # Summary by subsidiary
        if len(anom_df) > 0:
            sub_summary = anom_df['Subsidiary'].value_counts()
            print(f"\n📋 Anomaly summary by subsidiary:")
            for sub, count in sub_summary.items():
                print(f"   • {sub}: {count} anomalies")

            status_summary = anom_df['Status'].value_counts()
            print(f"\n🔍 Analysis status summary:")
            for status, count in status_summary.items():
                print(f"   • {status}: {count}")
    else:
        print(f"⚠️  No anomalies detected across all files")
        anom_df = pd.DataFrame(columns=[
            "Subsidiary","Account","Period","Pct Change","Abs Change (VND)",
            "Trigger(s)","Suggested likely cause","Status","Notes"
        ])

    # === WRITE TO WORKSHEET ===
    print(f"\n📝 Writing results to Excel worksheet...")
    row_count = 0
    for r in dataframe_to_rows(anom_df, index=False, header=True):
        ws.append(r)
        row_count += 1
    print(f"✅ Written {row_count} rows to worksheet (including header)")

    # === VISUAL FORMATTING ===
    print(f"\n🎨 Applying visual formatting to Excel output...")
    apply_excel_formatting_ws(ws, anom_df, CONFIG)
    print(f"✅ Excel formatting applied successfully")

    # === RETURN BYTES ===
    print(f"\n💾 Generating final Excel file...")
    bio = io.BytesIO()
    wb.save(bio)
    final_size = len(bio.getvalue())
    print(f"✅ Excel file generated successfully")
    print(f"   • Output size: {final_size:,} bytes ({final_size/1024:.1f} KB)")

    print(f"\n📊 Debug Files Summary:")
    print(f"   • Debug files created: {len(debug_files)}")
    for debug_name, debug_bytes in debug_files:
        print(f"     - {debug_name}: {len(debug_bytes):,} bytes ({len(debug_bytes)/1024:.1f} KB)")

    print(f"\n🎉 ===== AI VARIANCE ANALYSIS COMPLETED =====\n")
    return bio.getvalue(), debug_files

# -----------------------------------------------------------------------------
# Comprehensive Revenue Impact Analysis
# -----------------------------------------------------------------------------

def clean_numeric_value(val):
    """Convert value to numeric, handling various formats"""
    if pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(',', '').replace(' ', ''))
    except:
        return 0.0

def analyze_comprehensive_revenue_impact_from_bytes(xl_bytes: bytes, filename: str) -> dict:
    """
    Comprehensive Revenue Impact Analysis
    Answers specific questions:
    1. If revenue increases (511*), which specific revenue accounts drive the increase?
    2. Which customers/entities drive the revenue changes for each account?
    3. Gross margin analysis: (Revenue - Cost)/Revenue and risk identification
    4. Utility revenue vs cost pairing analysis
    """
    try:
        xls = pd.ExcelFile(io.BytesIO(xl_bytes))

        if 'PL Breakdown' not in xls.sheet_names:
            return {"error": "PL Breakdown sheet not found"}

        pl_df = pd.read_excel(xls, sheet_name='PL Breakdown')

        # Find data start row
        data_start_row = None
        for i, row in pl_df.iterrows():
            if str(row.iloc[1]).strip().lower() == 'entity':
                data_start_row = i
                break

        if data_start_row is None:
            return {"error": "Could not find data start row with 'Entity' header"}

        # Extract month columns
        month_headers = pl_df.iloc[data_start_row + 1].fillna('').astype(str).tolist()
        month_cols = []
        for h2 in month_headers:
            if any(month in str(h2) for month in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug']):
                month_cols.append(str(h2).strip())

        # Extract data
        data_df = pl_df.iloc[data_start_row + 2:].copy()
        actual_col_count = len(data_df.columns)
        new_columns = ['Account_Description', 'Entity', 'Account_Code']
        new_columns.extend(month_cols)
        while len(new_columns) < actual_col_count:
            new_columns.append(f'Extra_{len(new_columns)}')
        data_df.columns = new_columns[:actual_col_count]
        data_df = data_df.dropna(how='all')

        # Extract subsidiary name
        subsidiary = extract_subsidiary_name_from_bytes(xl_bytes, filename)

        analysis_result = {
            'subsidiary': subsidiary,
            'filename': filename,
            'months_analyzed': month_cols[:8],
            'total_revenue_analysis': {},
            'revenue_by_account': {},
            'gross_margin_analysis': {},
            'utility_analysis': {},
            'risk_assessment': []
        }

        # =====================================
        # 1. TOTAL REVENUE ANALYSIS (511*)
        # =====================================

        total_revenue_by_month = {}
        for month in month_cols[:8]:
            month_total = 0
            for i, row in data_df.iterrows():
                entity = str(row['Entity']) if 'Entity' in row and pd.notna(row['Entity']) else ''
                if entity and entity != 'nan' and not entity.startswith('Total'):
                    # Check if under a 511* revenue account
                    for prev_i in range(max(0, i-10), i):
                        if prev_i < len(data_df):
                            prev_desc = str(data_df.iloc[prev_i]['Account_Description']) if pd.notna(data_df.iloc[prev_i]['Account_Description']) else ''
                            if '511' in prev_desc and 'revenue' in prev_desc.lower():
                                val = clean_numeric_value(row[month])
                                month_total += val
                                break
            total_revenue_by_month[month] = month_total

        # Calculate month-over-month changes
        months = list(total_revenue_by_month.keys())
        total_revenue_changes = []
        for i in range(1, len(months)):
            prev_month = months[i-1]
            curr_month = months[i]
            prev_revenue = total_revenue_by_month[prev_month]
            curr_revenue = total_revenue_by_month[curr_month]
            change = curr_revenue - prev_revenue
            pct_change = (change / prev_revenue * 100) if prev_revenue != 0 else 0

            total_revenue_changes.append({
                'from': prev_month,
                'to': curr_month,
                'prev_value': prev_revenue,
                'curr_value': curr_revenue,
                'change': change,
                'pct_change': pct_change
            })

        analysis_result['total_revenue_analysis'] = {
            'monthly_totals': total_revenue_by_month,
            'changes': total_revenue_changes
        }

        # =====================================
        # 2. REVENUE BY ACCOUNT TYPE (511.xxx)
        # =====================================

        revenue_accounts = {}
        current_account = None

        for i, row in data_df.iterrows():
            account_desc = str(row['Account_Description']) if pd.notna(row['Account_Description']) else ''
            entity = str(row['Entity']) if pd.notna(row['Entity']) else ''

            # Revenue account headers
            if '511' in account_desc and 'revenue' in account_desc.lower():
                current_account = account_desc
                if current_account not in revenue_accounts:
                    revenue_accounts[current_account] = {
                        'entities': {},
                        'monthly_totals': {month: 0 for month in month_cols[:8]}
                    }

            # Entity data under current account
            elif current_account and entity and entity != 'nan' and not entity.startswith('Total'):
                if entity not in revenue_accounts[current_account]['entities']:
                    revenue_accounts[current_account]['entities'][entity] = {}

                for month in month_cols[:8]:
                    val = clean_numeric_value(row[month])
                    revenue_accounts[current_account]['entities'][entity][month] = val
                    revenue_accounts[current_account]['monthly_totals'][month] += val

        # Analyze each revenue account
        for account, data in revenue_accounts.items():
            months = list(data['monthly_totals'].keys())
            account_changes = []

            for i in range(1, len(months)):
                prev_month = months[i-1]
                curr_month = months[i]
                prev_val = data['monthly_totals'][prev_month]
                curr_val = data['monthly_totals'][curr_month]
                change = curr_val - prev_val
                pct_change = (change / prev_val * 100) if prev_val != 0 else 0

                account_changes.append({
                    'from': prev_month,
                    'to': curr_month,
                    'change': change,
                    'pct_change': pct_change,
                    'prev_val': prev_val,
                    'curr_val': curr_val
                })

            # Find biggest change for customer analysis
            biggest_change = max(account_changes, key=lambda x: abs(x['change'])) if account_changes else None
            customer_impacts = []

            if biggest_change and abs(biggest_change['change']) > 1000000:  # > 1M VND
                for entity, entity_data in data['entities'].items():
                    prev_val = entity_data.get(biggest_change['from'], 0)
                    curr_val = entity_data.get(biggest_change['to'], 0)
                    entity_change = curr_val - prev_val

                    if abs(entity_change) > 100000:  # > 100K VND
                        customer_impacts.append({
                            'entity': entity,
                            'change': entity_change,
                            'prev_val': prev_val,
                            'curr_val': curr_val,
                            'pct_change': (entity_change / prev_val * 100) if prev_val != 0 else 0
                        })

                customer_impacts.sort(key=lambda x: abs(x['change']), reverse=True)

            revenue_accounts[account]['changes'] = account_changes
            revenue_accounts[account]['biggest_change'] = biggest_change
            revenue_accounts[account]['customer_impacts'] = customer_impacts[:5]  # Top 5

        analysis_result['revenue_by_account'] = revenue_accounts

        # =====================================
        # 3. GROSS MARGIN ANALYSIS
        # =====================================

        cost_accounts = {}
        current_cost_account = None

        # Look for 632* cost accounts
        for i, row in data_df.iterrows():
            account_desc = str(row['Account_Description']) if pd.notna(row['Account_Description']) else ''
            entity = str(row['Entity']) if pd.notna(row['Entity']) else ''

            if '632' in account_desc and 'cost' in account_desc.lower():
                current_cost_account = account_desc
                if current_cost_account not in cost_accounts:
                    cost_accounts[current_cost_account] = {
                        'entities': {},
                        'monthly_totals': {month: 0 for month in month_cols[:8]}
                    }

            elif current_cost_account and entity and entity != 'nan' and not entity.startswith('Total'):
                if entity not in cost_accounts[current_cost_account]['entities']:
                    cost_accounts[current_cost_account]['entities'][entity] = {}

                for month in month_cols[:8]:
                    val = clean_numeric_value(row[month])
                    cost_accounts[current_cost_account]['entities'][entity][month] = val
                    cost_accounts[current_cost_account]['monthly_totals'][month] += val

        # Calculate gross margins
        gross_margin_trend = []
        for i in range(len(months)):
            month = months[i]
            total_revenue = total_revenue_by_month[month]
            total_cost = sum([cost_data['monthly_totals'][month] for cost_data in cost_accounts.values()])

            if total_revenue > 0:
                gross_margin_pct = ((total_revenue - total_cost) / total_revenue) * 100
                gross_margin_trend.append({
                    'month': month,
                    'revenue': total_revenue,
                    'cost': total_cost,
                    'gross_margin_pct': gross_margin_pct
                })

                if i > 0:
                    prev_gm = gross_margin_trend[i-1]['gross_margin_pct']
                    gm_change = gross_margin_pct - prev_gm

                    if abs(gm_change) > 1:  # > 1% change
                        risk_level = "HIGH" if gm_change < -2 else "MEDIUM"
                        analysis_result['risk_assessment'].append({
                            'type': 'Gross Margin Change',
                            'period': f"{gross_margin_trend[i-1]['month']} → {month}",
                            'change': gm_change,
                            'risk_level': risk_level,
                            'description': f"Gross margin changed by {gm_change:+.1f}%"
                        })

        analysis_result['gross_margin_analysis'] = {
            'trend': gross_margin_trend,
            'cost_accounts': cost_accounts
        }

        # =====================================
        # 4. UTILITY ANALYSIS
        # =====================================

        utility_revenue = None
        utility_cost = None

        for account, data in revenue_accounts.items():
            if 'utilit' in account.lower():
                utility_revenue = data
                break

        for account, data in cost_accounts.items():
            if 'utilit' in account.lower():
                utility_cost = data
                break

        if utility_revenue and utility_cost:
            utility_margins = []
            for month in months:
                rev = utility_revenue['monthly_totals'][month]
                cost = utility_cost['monthly_totals'][month]
                if rev > 0:
                    gm_pct = ((rev - cost) / rev) * 100
                    utility_margins.append({
                        'month': month,
                        'revenue': rev,
                        'cost': cost,
                        'margin_pct': gm_pct
                    })

            analysis_result['utility_analysis'] = {
                'available': True,
                'margins': utility_margins
            }
        else:
            analysis_result['utility_analysis'] = {
                'available': False,
                'reason': 'Could not find matching utility revenue/cost accounts'
            }

        # =====================================
        # 5. SUMMARY METRICS
        # =====================================

        analysis_result['summary'] = {
            'total_accounts': len([a for a in revenue_accounts if not a.startswith('Total')]),
            'highest_variance_account': max(revenue_accounts.items(),
                                          key=lambda x: max([abs(c['change']) for c in x[1].get('changes', [])], default=0))[0] if revenue_accounts else None,
            'total_revenue_latest': total_revenue_by_month[months[-1]] if months else 0,
            'gross_margin_latest': gross_margin_trend[-1]['gross_margin_pct'] if gross_margin_trend else 0,
            'risk_periods': [r for r in analysis_result['risk_assessment'] if r['risk_level'] == 'HIGH']
        }

        return analysis_result

    except Exception as e:
        return {"error": f"Comprehensive analysis failed: {str(e)}"}

def analyze_revenue_impact_from_bytes(xl_bytes: bytes, filename: str) -> dict:
    """
    Comprehensive revenue impact analysis from Excel bytes
    Returns structured data for frontend display
    """
    try:
        xls = pd.ExcelFile(io.BytesIO(xl_bytes))

        if 'PL Breakdown' not in xls.sheet_names:
            return {"error": "PL Breakdown sheet not found"}

        pl_df = pd.read_excel(xls, sheet_name='PL Breakdown')

        # Find data start row
        data_start_row = None
        for i, row in pl_df.iterrows():
            if str(row.iloc[1]).strip().lower() == 'entity':
                data_start_row = i
                break

        if data_start_row is None:
            return {"error": "Could not find data start row with 'Entity' header"}

        # Extract month columns
        month_headers = pl_df.iloc[data_start_row + 1].fillna('').astype(str).tolist()
        month_cols = []
        for h2 in month_headers:
            if any(month in str(h2) for month in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug']):
                month_cols.append(str(h2).strip())

        # Extract data
        data_df = pl_df.iloc[data_start_row + 2:].copy()
        actual_col_count = len(data_df.columns)
        new_columns = ['Account_Description', 'Entity', 'Account_Code']
        new_columns.extend(month_cols)
        while len(new_columns) < actual_col_count:
            new_columns.append(f'Extra_{len(new_columns)}')
        data_df.columns = new_columns[:actual_col_count]
        data_df = data_df.dropna(how='all')

        # Extract subsidiary name
        subsidiary = extract_subsidiary_name_from_bytes(xl_bytes, filename)

        # 1. Calculate total revenue by month
        total_revenue_by_month = {}
        for month in month_cols[:8]:
            month_total = 0
            for i, row in data_df.iterrows():
                entity = str(row['Entity']) if 'Entity' in row and pd.notna(row['Entity']) else ''
                if entity and entity != 'nan' and not entity.startswith('Total'):
                    # Check if under a 511* revenue account
                    for prev_i in range(max(0, i-10), i):
                        if prev_i < len(data_df):
                            prev_desc = str(data_df.iloc[prev_i]['Account_Description']) if pd.notna(data_df.iloc[prev_i]['Account_Description']) else ''
                            if '511' in prev_desc and 'revenue' in prev_desc.lower():
                                val = clean_numeric_value(row[month])
                                month_total += val
                                break
            total_revenue_by_month[month] = month_total

        # 2. Analyze revenue by account type
        revenue_accounts = {}
        current_account = None

        for i, row in data_df.iterrows():
            account_desc = str(row['Account_Description']) if pd.notna(row['Account_Description']) else ''
            entity = str(row['Entity']) if pd.notna(row['Entity']) else ''

            # Revenue account headers
            if '511' in account_desc and 'revenue' in account_desc.lower():
                current_account = account_desc
                if current_account not in revenue_accounts:
                    revenue_accounts[current_account] = {
                        'entities': {},
                        'monthly_totals': {month: 0 for month in month_cols[:8]}
                    }

            # Entity data under current account
            elif current_account and entity and entity != 'nan' and not entity.startswith('Total'):
                if entity not in revenue_accounts[current_account]['entities']:
                    revenue_accounts[current_account]['entities'][entity] = {}

                for month in month_cols[:8]:
                    val = clean_numeric_value(row[month])
                    revenue_accounts[current_account]['entities'][entity][month] = val
                    revenue_accounts[current_account]['monthly_totals'][month] += val

        # 3. Calculate changes and impacts
        months = list(total_revenue_by_month.keys())
        total_revenue_changes = []

        for i in range(1, len(months)):
            prev_month = months[i-1]
            curr_month = months[i]
            prev_val = total_revenue_by_month[prev_month]
            curr_val = total_revenue_by_month[curr_month]
            change = curr_val - prev_val
            pct_change = (change / prev_val * 100) if prev_val != 0 else 0

            total_revenue_changes.append({
                'from': prev_month,
                'to': curr_month,
                'prev_value': prev_val,
                'curr_value': curr_val,
                'change': change,
                'pct_change': pct_change
            })

        # 4. Account-level analysis with customer breakdowns
        account_analysis = []

        for account, data in revenue_accounts.items():
            if not account.startswith('Total'):  # Skip total rows
                months = list(data['monthly_totals'].keys())
                account_changes = []

                for i in range(1, len(months)):
                    prev_month = months[i-1]
                    curr_month = months[i]
                    prev_val = data['monthly_totals'][prev_month]
                    curr_val = data['monthly_totals'][curr_month]
                    change = curr_val - prev_val
                    pct_change = (change / prev_val * 100) if prev_val != 0 else 0

                    account_changes.append({
                        'from': prev_month,
                        'to': curr_month,
                        'change': change,
                        'pct_change': pct_change,
                        'prev_val': prev_val,
                        'curr_val': curr_val
                    })

                # Find biggest change for customer analysis
                biggest_change = max(account_changes, key=lambda x: abs(x['change'])) if account_changes else None
                customer_impacts = []

                if biggest_change and abs(biggest_change['change']) > 1000000:  # > 1M VND
                    for entity, entity_data in data['entities'].items():
                        prev_val = entity_data.get(biggest_change['from'], 0)
                        curr_val = entity_data.get(biggest_change['to'], 0)
                        entity_change = curr_val - prev_val

                        if abs(entity_change) > 100000:  # > 100K VND
                            customer_impacts.append({
                                'entity': entity,
                                'change': entity_change,
                                'prev_val': prev_val,
                                'curr_val': curr_val,
                                'pct_change': (entity_change / prev_val * 100) if prev_val != 0 else 0
                            })

                    customer_impacts.sort(key=lambda x: abs(x['change']), reverse=True)

                account_analysis.append({
                    'account': account,
                    'changes': account_changes,
                    'biggest_change': biggest_change,
                    'customer_impacts': customer_impacts[:5]  # Top 5
                })

        # 5. Risk analysis and gross margin calculation
        risk_periods = []

        # Simple gross margin estimation (would need cost data for full analysis)
        for change in total_revenue_changes:
            if abs(change['pct_change']) > 5:  # > 5% change
                risk_level = "HIGH" if abs(change['pct_change']) > 20 else "MEDIUM"
                risk_periods.append({
                    'period': f"{change['from']} → {change['to']}",
                    'change': change['change'],
                    'pct_change': change['pct_change'],
                    'risk_level': risk_level,
                    'description': f"Revenue changed by {change['pct_change']:+.1f}%"
                })

        return {
            "subsidiary": subsidiary,
            "months_analyzed": months,
            "total_revenue_trend": total_revenue_by_month,
            "total_revenue_changes": total_revenue_changes,
            "account_analysis": account_analysis,
            "risk_periods": risk_periods,
            "summary": {
                "total_accounts": len([a for a in account_analysis if not a['account'].startswith('Total')]),
                "highest_variance_account": max(account_analysis, key=lambda x: max([abs(c['change']) for c in x['changes']], default=0))['account'] if account_analysis else None,
                "total_revenue_latest": total_revenue_by_month[months[-1]] if months else 0
            }
        }

    except Exception as e:
        return {"error": f"Analysis failed: {str(e)}"}