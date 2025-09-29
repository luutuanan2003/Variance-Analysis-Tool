# app/services/recent_months_service.py
"""Recent months analysis service for focused two-month variance analysis."""

import io
import re
import asyncio
from typing import List, Tuple, Optional, Dict, Any
from datetime import datetime
from calendar import month_name
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from ..data.data_utils import normalize_period_label, month_key, DEFAULT_CONFIG
from ..data.excel_processing import process_financial_tab_from_bytes, extract_subsidiary_name_from_bytes
from ..analysis.anomaly_detection import build_anoms_python_mode
from ..utils.logging_config import get_logger

logger = get_logger(__name__)

class RecentMonthsAnalysisService:
    """Service for performing recent months analysis focusing on current and previous month."""

    def __init__(self):
        self.logger = logger

    def extract_target_month_from_bs_breakdown(self, xl_bytes: bytes) -> Optional[str]:
        """
        Extract target month from row 4 of BS Breakdown sheet.

        Expected format in row 4: "End of Aug 2025"
        Returns: "Aug 2025" or None if not found
        """
        try:
            # Load the BS Breakdown sheet to examine row 4
            wb = load_workbook(io.BytesIO(xl_bytes), read_only=True, data_only=True)

            if "BS Breakdown" not in wb.sheetnames:
                self.logger.warning("BS Breakdown sheet not found")
                return None

            sheet = wb["BS Breakdown"]

            # Check row 4 for the target month declaration
            # Row 4 in openpyxl is index 4 (1-based indexing)
            for col in range(1, 20):  # Check first 20 columns
                cell_value = sheet.cell(row=4, column=col).value
                if cell_value and isinstance(cell_value, str):
                    # Look for patterns like "End of Aug 2025", "As of Aug 2025", etc.
                    match = re.search(r'(?:end\s+of|as\s+of|tinh\s+den|tính\s+đến)\s+(\w+\s+\d{4})',
                                    str(cell_value), re.I)
                    if match:
                        period_str = match.group(1)
                        normalized = normalize_period_label(period_str)
                        self.logger.info(f"Found target month in BS Breakdown row 4: {normalized}")
                        return normalized

            self.logger.warning("Target month not found in BS Breakdown row 4")
            return None

        except Exception as e:
            self.logger.error(f"Error extracting target month: {e}", exc_info=True)
            return None

    def calculate_recent_months(self, target_month: str) -> Tuple[str, str]:
        """
        Calculate current and previous month from target month.

        Args:
            target_month: e.g., "Aug 2025"

        Returns:
            Tuple of (current_month, previous_month) e.g., ("Aug 2025", "Jul 2025")
        """
        try:
            # Parse the target month
            match = re.match(r'(\w+)\s+(\d{4})', target_month.strip())
            if not match:
                raise ValueError(f"Invalid month format: {target_month}")

            month_name_str = match.group(1).lower()
            year = int(match.group(2))

            # Map month names to numbers
            month_map = {
                'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
            }

            if month_name_str not in month_map:
                raise ValueError(f"Unknown month: {month_name_str}")

            current_month_num = month_map[month_name_str]

            # Calculate previous month
            if current_month_num == 1:
                prev_month_num = 12
                prev_year = year - 1
            else:
                prev_month_num = current_month_num - 1
                prev_year = year

            # Convert back to month names
            month_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

            current_month = f"{month_names[current_month_num]} {year}"
            previous_month = f"{month_names[prev_month_num]} {prev_year}"

            self.logger.info(f"Recent months calculated: {previous_month} → {current_month}")
            return current_month, previous_month

        except Exception as e:
            self.logger.error(f"Error calculating recent months: {e}", exc_info=True)
            raise ValueError(f"Failed to calculate recent months from {target_month}: {str(e)}")

    def filter_dataframe_to_recent_months(self, df: pd.DataFrame, current_month: str, previous_month: str) -> pd.DataFrame:
        """
        Filter DataFrame to only include the recent two months.

        Args:
            df: Financial data DataFrame with month columns
            current_month: e.g., "Aug 2025"
            previous_month: e.g., "Jul 2025"

        Returns:
            Filtered DataFrame with only recent months + non-month columns
        """
        if df.empty:
            return df

        # Identify month columns
        month_cols = []
        for col in df.columns:
            normalized = normalize_period_label(str(col))
            if re.match(r'^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+\d{4}$', normalized, re.I):
                month_cols.append(col)

        # Find the specific months we want
        target_cols = []
        for col in month_cols:
            normalized = normalize_period_label(str(col))
            if normalized.lower() == current_month.lower() or normalized.lower() == previous_month.lower():
                target_cols.append(col)

        # Keep non-month columns and the two target months
        non_month_cols = [col for col in df.columns if col not in month_cols]
        filtered_cols = non_month_cols + sorted(target_cols, key=month_key)

        self.logger.info(f"Filtered to columns: {filtered_cols}")
        return df[filtered_cols].copy()

    def _create_simple_variance_analysis(self, bs_df: pd.DataFrame, current_month: str, previous_month: str, subsidiary: str) -> pd.DataFrame:
        """
        Create a simple variance analysis for the recent two months.

        Args:
            bs_df: Filtered balance sheet DataFrame with two months
            current_month: Current month name (e.g., "Aug 2025")
            previous_month: Previous month name (e.g., "Jul 2025")
            subsidiary: Subsidiary name

        Returns:
            DataFrame with variance analysis results
        """
        try:
            if bs_df.empty or len(bs_df.columns) < 4:
                return pd.DataFrame()

            # Find the month columns in the filtered data
            month_cols = []
            for col in bs_df.columns:
                normalized = normalize_period_label(str(col))
                if normalized.lower() in [current_month.lower(), previous_month.lower()]:
                    month_cols.append(col)

            if len(month_cols) < 2:
                self.logger.warning("Insufficient month columns for variance analysis")
                return pd.DataFrame()

            # Sort months chronologically (previous, current)
            month_cols.sort(key=month_key)
            prev_col, curr_col = month_cols[0], month_cols[1]

            # Create variance analysis
            variance_results = []

            for _, row in bs_df.iterrows():
                account_code = row.get('Account Code', '')
                account_name = row.get('Account Name', '')

                try:
                    # Get values and convert to float
                    prev_val = pd.to_numeric(row[prev_col], errors='coerce')
                    curr_val = pd.to_numeric(row[curr_col], errors='coerce')

                    # Skip if either value is NaN or both are zero
                    if pd.isna(prev_val) or pd.isna(curr_val) or (prev_val == 0 and curr_val == 0):
                        continue

                    # Calculate variance
                    delta = curr_val - prev_val
                    pct_change = (delta / prev_val * 100) if prev_val != 0 else float('inf')

                    # Only include significant changes (> 10% or absolute change > 1M)
                    if abs(pct_change) > 10 or abs(delta) > 1000000:
                        variance_results.append({
                            'Subsidiary': subsidiary,
                            'Account': f"{account_code} - {account_name}",
                            'Period': f"{previous_month} → {current_month}",
                            'Pct Change': f"{pct_change:.1f}%" if pct_change != float('inf') else "New",
                            'Abs Change (VND)': f"{delta:,.0f}",
                            'Trigger(s)': "Recent Months Variance",
                            'Suggested likely cause': self._suggest_cause_for_account(account_name, pct_change),
                            'Status': "Identified",
                            'Notes': f"Change from {prev_val:,.0f} to {curr_val:,.0f}"
                        })

                except Exception as e:
                    self.logger.debug(f"Error processing account {account_code}: {e}")
                    continue

            if variance_results:
                return pd.DataFrame(variance_results)
            else:
                return pd.DataFrame(columns=[
                    "Subsidiary", "Account", "Period", "Pct Change", "Abs Change (VND)",
                    "Trigger(s)", "Suggested likely cause", "Status", "Notes"
                ])

        except Exception as e:
            self.logger.error(f"Error in simple variance analysis: {e}", exc_info=True)
            return pd.DataFrame()

    def _suggest_cause_for_account(self, account_name: str, pct_change: float) -> str:
        """Suggest likely cause based on account name and change magnitude."""
        account_lower = account_name.lower()

        if 'revenue' in account_lower or 'sales' in account_lower:
            return "Revenue fluctuation - check customer activity"
        elif 'cash' in account_lower or 'bank' in account_lower:
            return "Cash flow changes - review collections/payments"
        elif 'receivable' in account_lower:
            return "Receivables change - verify customer payments"
        elif 'inventory' in account_lower:
            return "Inventory movement - check purchases/sales"
        elif 'payable' in account_lower:
            return "Payables change - review vendor payments"
        elif abs(pct_change) > 50:
            return "Significant change - requires detailed investigation"
        elif abs(pct_change) > 25:
            return "Material change - review supporting documentation"
        else:
            return "Normal business variance - monitor for trends"

    def analyze_recent_months_sync(
        self,
        files: List[Tuple[str, bytes]],
        config: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        Completely synchronous version of recent months analysis for integration with processing pipeline.
        """
        try:
            self.logger.info("Starting synchronous recent months analysis")

            if not files:
                raise ValueError("No files provided for analysis")

            # Use default config if none provided
            analysis_config = config or DEFAULT_CONFIG.copy()

            # Process each file
            results = []

            for filename, file_bytes in files:
                self.logger.info(f"Processing file: {filename}")

                # Extract subsidiary name
                subsidiary = extract_subsidiary_name_from_bytes(file_bytes, filename)

                # Extract target month from BS Breakdown row 4
                target_month = self.extract_target_month_from_bs_breakdown(file_bytes)
                if not target_month:
                    raise ValueError(f"Could not extract target month from {filename}")

                # Calculate recent months
                current_month, previous_month = self.calculate_recent_months(target_month)

                # Process BS Breakdown sheet
                bs_df, bs_cols = process_financial_tab_from_bytes(file_bytes, "BS Breakdown", "BS", subsidiary)
                bs_filtered = self.filter_dataframe_to_recent_months(bs_df, current_month, previous_month)

                # Process PL Breakdown sheet
                try:
                    pl_df, pl_cols = process_financial_tab_from_bytes(file_bytes, "PL Breakdown", "PL", subsidiary)
                    pl_filtered = self.filter_dataframe_to_recent_months(pl_df, current_month, previous_month)
                except Exception as e:
                    self.logger.warning(f"Could not process PL Breakdown: {e}")
                    pl_filtered = pd.DataFrame()

                # For recent months analysis, we'll skip complex anomaly detection
                # since we only have 2 months of data which isn't enough for reliable anomaly detection
                # Instead, we'll create a simple variance analysis
                anomalies = self._create_simple_variance_analysis(bs_filtered, current_month, previous_month, subsidiary)

                results.append({
                    'filename': filename,
                    'subsidiary': subsidiary,
                    'target_month': target_month,
                    'current_month': current_month,
                    'previous_month': previous_month,
                    'bs_data': bs_filtered,
                    'pl_data': pl_filtered,
                    'anomalies': anomalies
                })

            # Create output Excel file
            output_bytes = self._create_recent_months_excel_sync(results, analysis_config)

            self.logger.info("Synchronous recent months analysis completed successfully")
            return output_bytes

        except Exception as e:
            self.logger.error(f"Synchronous recent months analysis failed: {e}", exc_info=True)
            raise

    def _create_recent_months_excel_sync(self, results: List[Dict], config: Dict[str, Any]) -> bytes:
        """Create Excel file with Recent Months Analysis sheet (synchronous version)."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Recent Months Analysis"

        # Header styling
        from openpyxl.styles import PatternFill, Font, Border, Side

        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_font = Font(bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        row = 1

        # Title
        ws[f"A{row}"] = "Recent Months Analysis Report"
        ws[f"A{row}"].font = Font(bold=True, size=16)
        row += 2

        # Summary section
        ws[f"A{row}"] = "Analysis Summary"
        ws[f"A{row}"].font = section_font
        row += 1

        # Add summary information
        for result in results:
            ws[f"A{row}"] = "File:"
            ws[f"B{row}"] = result['filename']
            row += 1

            ws[f"A{row}"] = "Subsidiary:"
            ws[f"B{row}"] = result['subsidiary']
            row += 1

            ws[f"A{row}"] = "Analysis Period:"
            ws[f"B{row}"] = f"{result['previous_month']} → {result['current_month']}"
            row += 1

            ws[f"A{row}"] = "Target Month:"
            ws[f"B{row}"] = result['target_month']
            row += 2

        # Anomalies section
        ws[f"A{row}"] = "Identified Anomalies"
        ws[f"A{row}"].font = section_font
        row += 1

        # Headers for anomalies table
        headers = ["Subsidiary", "Account", "Period", "Pct Change", "Abs Change (VND)", "Trigger(s)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        # Add anomalies data
        anomaly_count = 0
        for result in results:
            anomalies = result['anomalies']
            if isinstance(anomalies, pd.DataFrame) and not anomalies.empty:
                for _, anomaly in anomalies.iterrows():
                    anomaly_count += 1
                    ws[f"A{row}"] = anomaly.get('Subsidiary', result['subsidiary'])
                    ws[f"B{row}"] = anomaly.get('Account', '')
                    ws[f"C{row}"] = anomaly.get('Period', '')
                    ws[f"D{row}"] = anomaly.get('Pct Change', '')
                    ws[f"E{row}"] = anomaly.get('Abs Change (VND)', '')
                    ws[f"F{row}"] = anomaly.get('Trigger(s)', '')

                    # Apply borders
                    for col in range(1, 7):
                        ws.cell(row=row, column=col).border = thin_border
                    row += 1

        if anomaly_count == 0:
            ws[f"A{row}"] = "No anomalies detected for the recent months period."
            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    async def analyze_recent_months(
        self,
        files: List[Tuple[str, bytes]],
        config: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        Perform recent months analysis on uploaded files.

        Args:
            files: List of (filename, file_bytes) tuples
            config: Analysis configuration

        Returns:
            Excel file bytes with Recent Months Analysis sheet
        """
        try:
            self.logger.info("Starting recent months analysis")

            if not files:
                raise ValueError("No files provided for analysis")

            # Use default config if none provided
            analysis_config = config or DEFAULT_CONFIG.copy()

            # Process each file
            results = []

            for filename, file_bytes in files:
                self.logger.info(f"Processing file: {filename}")

                # Extract subsidiary name
                subsidiary = extract_subsidiary_name_from_bytes(file_bytes, filename)

                # Extract target month from BS Breakdown row 4
                target_month = self.extract_target_month_from_bs_breakdown(file_bytes)
                if not target_month:
                    raise ValueError(f"Could not extract target month from {filename}")

                # Calculate recent months
                current_month, previous_month = self.calculate_recent_months(target_month)

                # Process BS Breakdown sheet
                bs_df, bs_cols = process_financial_tab_from_bytes(file_bytes, "BS Breakdown", "BS", subsidiary)
                bs_filtered = self.filter_dataframe_to_recent_months(bs_df, current_month, previous_month)

                # Process PL Breakdown sheet
                try:
                    pl_df, pl_cols = process_financial_tab_from_bytes(file_bytes, "PL Breakdown", "PL", subsidiary)
                    pl_filtered = self.filter_dataframe_to_recent_months(pl_df, current_month, previous_month)
                except Exception as e:
                    self.logger.warning(f"Could not process PL Breakdown: {e}")
                    pl_filtered = pd.DataFrame()

                # Run anomaly detection on filtered data
                if not bs_filtered.empty:
                    anomalies = build_anoms_python_mode(
                        bs_df=bs_filtered,
                        pl_df=pl_filtered,
                        subsidiary=subsidiary,
                        CONFIG=analysis_config
                    )
                else:
                    anomalies = []

                results.append({
                    'filename': filename,
                    'subsidiary': subsidiary,
                    'target_month': target_month,
                    'current_month': current_month,
                    'previous_month': previous_month,
                    'bs_data': bs_filtered,
                    'pl_data': pl_filtered,
                    'anomalies': anomalies
                })

            # Create output Excel file
            output_bytes = self._create_recent_months_excel(results, analysis_config)

            self.logger.info("Recent months analysis completed successfully")
            return output_bytes

        except Exception as e:
            self.logger.error(f"Recent months analysis failed: {e}", exc_info=True)
            raise

    def _create_recent_months_excel(self, results: List[Dict], config: Dict[str, Any]) -> bytes:
        """Create Excel file with Recent Months Analysis sheet."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Recent Months Analysis"

        # Header styling
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_font = Font(bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        row = 1

        # Title
        ws[f"A{row}"] = "Recent Months Analysis Report"
        ws[f"A{row}"].font = Font(bold=True, size=16)
        row += 2

        # Summary section
        ws[f"A{row}"] = "Analysis Summary"
        ws[f"A{row}"].font = section_font
        row += 1

        # Add summary information
        for result in results:
            ws[f"A{row}"] = "File:"
            ws[f"B{row}"] = result['filename']
            row += 1

            ws[f"A{row}"] = "Subsidiary:"
            ws[f"B{row}"] = result['subsidiary']
            row += 1

            ws[f"A{row}"] = "Analysis Period:"
            ws[f"B{row}"] = f"{result['previous_month']} → {result['current_month']}"
            row += 1

            ws[f"A{row}"] = "Target Month:"
            ws[f"B{row}"] = result['target_month']
            row += 2

        # Anomalies section
        ws[f"A{row}"] = "Identified Anomalies"
        ws[f"A{row}"].font = section_font
        row += 1

        # Headers for anomalies table
        headers = ["Rule", "Entity", "Account", "Month", "Value", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        # Add anomalies data
        anomaly_count = 0
        for result in results:
            for anomaly in result['anomalies']:
                anomaly_count += 1
                ws[f"A{row}"] = anomaly.get('Rule', '')
                ws[f"B{row}"] = anomaly.get('Entity', '')
                ws[f"C{row}"] = anomaly.get('Account', '')
                ws[f"D{row}"] = anomaly.get('Month', '')
                ws[f"E{row}"] = anomaly.get('Value', '')
                ws[f"F{row}"] = anomaly.get('Description', '')

                # Apply borders
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = thin_border
                row += 1

        if anomaly_count == 0:
            ws[f"A{row}"] = "No anomalies detected for the recent months period."
            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

# Create singleton instance
recent_months_service = RecentMonthsAnalysisService()